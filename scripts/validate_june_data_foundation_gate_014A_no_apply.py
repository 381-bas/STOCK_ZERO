#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
014A_JUNE_DATA_FOUNDATION_GATE_NO_APPLY

Validador local/read-only para perfilar base junio CONTROL_GESTION.

NO hace:
- Supabase writes
- SQL apply
- DDL
- loaders productivos
- refresh productivo
- movimiento de data
- git add / commit / stash apply
"""

from __future__ import annotations

import argparse
import calendar
import hashlib
import json
import math
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


PHASE_ID = "014A_JUNE_DATA_FOUNDATION_GATE_NO_APPLY"

EXPECTED_RUTA_ROWS = {
    "RUTA_RUTEROS_JUNIO_S1.xlsx": 3470,
    "RUTA_RUTEROS_JUNIO_S2.xlsx": 3543,
    "RUTA_RUTEROS_JUNIO_S3.xlsx": 3607,
    "RUTA_RUTEROS_JUNIO_S4.xlsx": 4057,
}

EXPECTED_KPIONE2_ROWS = 45737

MONTH_YEAR = 2026
MONTH_NUM = 6

NULLISH = {"", "N/A", "NA", "NULL", "NONE", "-", "SIN INFORMACION", "SIN INFORMACIÓN"}

CRITICALS = {
    "ruta_rutero": {
        "cod_rt": ["COD KPI ONE", "COD_RT", "COD RT", "CODIGO LOCAL", "CÓDIGO LOCAL", "COD LOCAL", "COD B2B"],
        "local": ["LOCAL", "NOMBRE LOCAL", "NOMBRE_LOCAL"],
        "cliente": ["CLIENTE", "MARCA"],
        "gestor": ["GESTORES", "GESTOR"],
        "rutero": ["RUTERO"],
        "reponedor": ["REPONEDOR"],
        "supervisor": ["SUPERVISOR"],
        "modalidad": ["MODALIDAD"],
        "lunes": ["LUNES"],
        "martes": ["MARTES"],
        "miercoles": ["MIERCOLES", "MIÉRCOLES"],
        "jueves": ["JUEVES"],
        "viernes": ["VIERNES"],
        "sabado": ["SABADO", "SÁBADO"],
        "domingo": ["DOMINGO"],
    },
    "kpione2": {
        "id": ["ID"],
        "sp_item_id": ["SP ITEM ID"],
        "cod_local": ["CODIGO LOCAL", "CÓDIGO LOCAL", "CODLOCAL", "COD RT", "COD_RT"],
        "cliente": ["MARCA", "CLIENTE"],
        "local": ["LOCAL", "NOMBRE LOCAL"],
        "reponedor": ["REPONEDOR", "TRABAJADOR"],
        "fecha": ["FECHA", "FECHA VISITA", "FECHA_REG", "FECHA REG"],
        "hora": ["HORA"],
        "n_fotos": ["N FOTOS", "CANT FOTOS", "CANT_FOTOS"],
        "link_foto": ["LINK FOTO", "LINK", "FOTO", "URL"],
        "visita": ["VISITA"],
        "registro_fuera_cruce": ["REGISTRO_FUERA_CRUCE", "REGISTRO FUERA CRUCE"],
        "semana": ["SEMANA", "SEM"],
    },
    "photo_admin": {
        "cod_local": ["CODIGO LOCAL", "CÓDIGO LOCAL", "CODLOCAL", "COD RT", "COD_RT", "SP ITEM ID"],
        "cliente": ["MARCA", "CLIENTE"],
        "local": ["LOCAL", "NOMBRE LOCAL"],
        "reponedor": ["REPONEDOR", "TRABAJADOR"],
        "fecha": ["FECHA", "FECHA VISITA", "CREADO"],
        "hora": ["HORA"],
        "n_fotos": ["N FOTOS", "CANT FOTOS", "CANT_FOTOS"],
        "link_foto": ["LINK FOTO", "LINK", "FOTO", "URL"],
    },
    "power_app": {
        "cliente": ["MARCA: TITULO", "MARCA: TÍTULO", "MARCA", "CLIENTE"],
        "local": ["LOCAL: TITLE", "LOCAL: LOCAL", "LOCAL"],
        "fecha": ["FECHA", "CREADO"],
        "registro_fuera_cruce": ["REGISTRO_FUERA_CRUCE", "REGISTRO FUERA CRUCE"],
        "semana": ["SEMANA", "SEM"],
    },
}

DAY_FIELDS = ["lunes", "martes", "miercoles", "jueves", "viernes", "sabado", "domingo"]


def norm(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().upper()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("_", " ").replace(".", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def is_nullish(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str):
        return norm(value) in NULLISH
    return False


def parse_date(value: Any) -> Optional[date]:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if 20000 <= float(value) <= 70000:
            try:
                parsed = from_excel(float(value))
                if isinstance(parsed, datetime):
                    return parsed.date()
                if isinstance(parsed, date):
                    return parsed
            except Exception:
                return None
        return None

    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None

        text = text.replace("/", "-").replace(".", "-")
        text = re.split(r"\s+", text)[0]

        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d-%m-%y", "%Y%m%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                pass

    return None


def week_start(d: date) -> date:
    return d - timedelta(days=d.weekday())


def operational_weeks_june_2026() -> list[dict[str, Any]]:
    first = date(2026, 6, 1)
    last = date(2026, 6, 30)

    start = first - timedelta(days=first.weekday())
    end = last + timedelta(days=6 - last.weekday())

    weeks = []
    cur = start

    while cur <= end:
        days = [cur + timedelta(days=i) for i in range(7)]
        days_in_month = [d for d in days if d.year == 2026 and d.month == 6]

        if len(days_in_month) >= 4:
            weeks.append({
                "label": f"S{len(weeks) + 1}",
                "week_start": cur.isoformat(),
                "week_end": (cur + timedelta(days=6)).isoformat(),
                "days_in_month": len(days_in_month),
            })

        cur += timedelta(days=7)

    return weeks


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def file_meta(path: Path) -> dict[str, Any]:
    stat = path.stat()
    return {
        "name": path.name,
        "path": str(path),
        "size_bytes": stat.st_size,
        "mtime_local": datetime.fromtimestamp(stat.st_mtime).astimezone().isoformat(timespec="seconds"),
        "sha256": sha256_file(path),
    }


def choose_sheet(wb, preferred: list[str]) -> str:
    available = wb.sheetnames
    by_norm = {norm(s): s for s in available}

    for wanted in preferred:
        if norm(wanted) in by_norm:
            return by_norm[norm(wanted)]

    return available[0]


def detect_header(ws, criticals: dict[str, list[str]], scan_rows: int = 25) -> tuple[int, list[Any]]:
    expected = {norm(x) for alternatives in criticals.values() for x in alternatives}

    best_row_idx = 1
    best_values = []
    best_score = -1
    best_nonempty = -1

    max_row = min(ws.max_row or 1, scan_rows)

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, values_only=True), start=1):
        values = list(row)
        normalized = [norm(v) for v in values if not is_nullish(v)]
        score = sum(1 for v in normalized if v in expected)
        nonempty = len(normalized)

        if score > best_score or (score == best_score and nonempty > best_nonempty):
            best_row_idx = row_idx
            best_values = values
            best_score = score
            best_nonempty = nonempty

    return best_row_idx, best_values


def resolve_fields(headers: list[Any], criticals: dict[str, list[str]]) -> dict[str, Any]:
    header_map = {}

    for idx, header in enumerate(headers):
        key = norm(header)
        if key and key not in header_map:
            header_map[key] = idx

    resolved = {}

    for field, alternatives in criticals.items():
        hit = None
        for alt in alternatives:
            key = norm(alt)
            if key in header_map:
                hit = {
                    "header": str(headers[header_map[key]]),
                    "index": header_map[key],
                }
                break
        resolved[field] = hit

    return resolved


def truthy_plan(value: Any) -> bool:
    if is_nullish(value):
        return False

    if isinstance(value, bool):
        return value

    if isinstance(value, (int, float)):
        return value != 0

    return norm(value) in {"1", "SI", "S", "TRUE", "VERDADERO", "X", "OK", "PLAN", "PLANIFICADO"}


def scan_sheet(ws, header_row: int, resolved: dict[str, Any], source_type: str, max_rows: Optional[int]) -> dict[str, Any]:
    data_rows = max((ws.max_row or 0) - header_row, 0)
    scan_limit = data_rows if max_rows is None else min(data_rows, max_rows)

    null_counts = {k: 0 for k, v in resolved.items() if v is not None}
    nonnull_counts = {k: 0 for k, v in resolved.items() if v is not None}
    distinct_values = {
        k: set()
        for k, v in resolved.items()
        if v is not None and k in {"cod_rt", "cod_local", "cliente", "local", "gestor", "rutero", "reponedor", "modalidad"}
    }

    date_profiles_raw = defaultdict(list)
    daily_counts = Counter()
    week_counts = Counter()
    plan_day_counts = Counter()

    for row in ws.iter_rows(min_row=header_row + 1, max_row=header_row + scan_limit, values_only=True):
        for field, meta in resolved.items():
            if meta is None:
                continue

            idx = meta["index"]
            value = row[idx] if idx < len(row) else None

            if is_nullish(value):
                null_counts[field] += 1
            else:
                nonnull_counts[field] += 1

                if field in distinct_values:
                    distinct_values[field].add(norm(value))

            if field == "fecha":
                d = parse_date(value)
                if d:
                    date_profiles_raw[field].append(d)
                    daily_counts[d.isoformat()] += 1
                    week_counts[week_start(d).isoformat()] += 1

            if source_type == "ruta_rutero" and field in DAY_FIELDS and truthy_plan(value):
                plan_day_counts[field] += 1

    date_profiles = {}

    for field, values in date_profiles_raw.items():
        if values:
            date_profiles[field] = {
                "parsed_count": len(values),
                "min_date": min(values).isoformat(),
                "max_date": max(values).isoformat(),
                "distinct_dates": len(set(values)),
            }
        else:
            date_profiles[field] = {"parsed_count": 0}

    return {
        "scanned_rows": scan_limit,
        "null_counts": null_counts,
        "nonnull_counts": nonnull_counts,
        "null_rates": {k: round(v / scan_limit, 6) if scan_limit else None for k, v in null_counts.items()},
        "distinct_counts": {k: len(v) for k, v in distinct_values.items()},
        "date_profiles": date_profiles,
        "daily_counts": dict(sorted(daily_counts.items())),
        "week_counts": dict(sorted(week_counts.items())),
        "ruta_plan_day_counts": dict(sorted(plan_day_counts.items())),
    }


def profile_excel(path: Path, source_type: str, preferred_sheets: list[str], criticals: dict[str, list[str]], max_rows: Optional[int]) -> dict[str, Any]:
    profile = {
        "source_type": source_type,
        "file": file_meta(path),
        "status": "ok",
    }

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        sheet_name = choose_sheet(wb, preferred_sheets)
        ws = wb[sheet_name]

        header_row, headers = detect_header(ws, criticals)
        resolved = resolve_fields(headers, criticals)
        missing = [k for k, v in resolved.items() if v is None]

        duplicated_headers = [
            {"header_norm": k, "count": v}
            for k, v in Counter(norm(h) for h in headers if not is_nullish(h)).items()
            if v > 1
        ]

        profile.update({
            "workbook_sheets": wb.sheetnames,
            "sheet": sheet_name,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "header_row": header_row,
            "data_rows": max((ws.max_row or 0) - header_row, 0),
            "headers": [None if h is None else str(h) for h in headers],
            "duplicate_headers": duplicated_headers,
            "blank_header_count": sum(1 for h in headers if is_nullish(h)),
            "resolved_fields": resolved,
            "missing_critical_fields": missing,
        })

        profile.update(scan_sheet(ws, header_row, resolved, source_type, max_rows))

        try:
            wb.close()
        except Exception:
            pass

    except Exception as exc:
        profile["status"] = "error"
        profile["error"] = repr(exc)

    return profile


def find_inputs(base: Path) -> dict[str, Any]:
    data = base / "data"

    ruta_dir = data / "RUTA_RUTERO" / "06 - JUNIO"
    ruta_files = sorted(ruta_dir.glob("RUTA_RUTEROS_JUNIO_S*.xlsx")) if ruta_dir.exists() else []

    photo_files = sorted(data.glob("photo-excel-admin_*.xlsx")) if data.exists() else []

    cumplimiento = data / "CUMPLIMIENTO_FRECUENCIA.xlsx"

    return {
        "base": str(base),
        "ruta_dir": str(ruta_dir),
        "ruta_files": [str(p) for p in ruta_files],
        "photo_files": [str(p) for p in photo_files],
        "cumplimiento_file": str(cumplimiento) if cumplimiento.exists() else None,
    }


def june_days() -> list[str]:
    return [
        date(2026, 6, day).isoformat()
        for day in range(1, calendar.monthrange(2026, 6)[1] + 1)
    ]


def coverage(profiles: list[dict[str, Any]], weeks: list[dict[str, Any]]) -> dict[str, Any]:
    month_days = june_days()
    week_starts = [w["week_start"] for w in weeks]

    out = {}

    for p in profiles:
        if p.get("status") != "ok":
            continue

        source = p["source_type"]

        if source in {"ruta_rutero", "photo_admin"}:
            label = f"{source}::{p['file']['name']}"
        else:
            label = source

        daily = p.get("daily_counts", {})
        weekly = p.get("week_counts", {})

        present_days = [d for d in month_days if daily.get(d, 0) > 0]
        missing_days = [d for d in month_days if daily.get(d, 0) == 0]

        present_weeks = [w for w in week_starts if weekly.get(w, 0) > 0]
        missing_weeks = [w for w in week_starts if weekly.get(w, 0) == 0]

        out[label] = {
            "month_days_with_records": len(present_days),
            "month_missing_days": missing_days,
            "operational_weeks_with_records": present_weeks,
            "operational_weeks_missing_records": missing_weeks,
            "total_month_records_by_date_field": sum(daily.get(d, 0) for d in month_days),
        }

    return out


def verdict(report: dict[str, Any]) -> dict[str, Any]:
    blockers = []
    warnings = []

    inv = report["input_inventory"]
    ruta_names = [Path(p).name for p in inv["ruta_files"]]

    missing_ruta = [name for name in EXPECTED_RUTA_ROWS if name not in ruta_names]
    if missing_ruta:
        blockers.append({"id": "MISSING_RUTA_JUNE_FILES", "detail": missing_ruta})

    if not inv["cumplimiento_file"]:
        blockers.append({"id": "MISSING_CUMPLIMIENTO_FRECUENCIA", "detail": "No existe data/CUMPLIMIENTO_FRECUENCIA.xlsx"})

    if len(inv["photo_files"]) == 0:
        warnings.append({"id": "NO_PHOTO_ADMIN_FILES", "detail": "No se encontraron data/photo-excel-admin_*.xlsx"})

    for p in report["profiles"]:
        if p.get("status") != "ok":
            blockers.append({"id": "PROFILE_ERROR", "file": p.get("file", {}).get("name"), "detail": p.get("error")})
            continue

        source = p["source_type"]
        missing = set(p.get("missing_critical_fields", []))

        if source == "ruta_rutero":
            fatal = sorted(missing.intersection({"cod_rt", "cliente", "local"}))
            if fatal:
                blockers.append({"id": "RUTA_CRITICAL_HEADERS_MISSING", "file": p["file"]["name"], "detail": fatal})

            expected = EXPECTED_RUTA_ROWS.get(p["file"]["name"])
            if expected is not None and p["data_rows"] != expected:
                warnings.append({
                    "id": "RUTA_ROW_COUNT_DIFF",
                    "file": p["file"]["name"],
                    "expected": expected,
                    "observed": p["data_rows"],
                    "delta": p["data_rows"] - expected,
                })

        if source == "kpione2":
            fatal = sorted(missing.intersection({"cod_local", "cliente", "local", "fecha"}))
            if fatal:
                blockers.append({"id": "KPIONE2_CRITICAL_HEADERS_MISSING", "file": p["file"]["name"], "detail": fatal})

            if p["data_rows"] != EXPECTED_KPIONE2_ROWS:
                warnings.append({
                    "id": "KPIONE2_ROW_COUNT_DIFF_VS_PHASE_INVENTORY",
                    "expected_from_user_inventory": EXPECTED_KPIONE2_ROWS,
                    "observed": p["data_rows"],
                    "delta": p["data_rows"] - EXPECTED_KPIONE2_ROWS,
                })

    cov = report["coverage"]
    kp = cov.get("kpione2")

    if kp:
        if kp["month_days_with_records"] == 0:
            blockers.append({
                "id": "KPIONE2_NO_JUNE_DATE_COVERAGE",
                "detail": "No hay fechas parseables de junio en DB (KPIONE2.0)",
            })
        elif kp["month_days_with_records"] < 28:
            warnings.append({
                "id": "KPIONE2_DAILY_GAPS_IN_JUNE",
                "days_with_records": kp["month_days_with_records"],
                "missing_days": kp["month_missing_days"],
            })

        if kp["operational_weeks_missing_records"]:
            warnings.append({
                "id": "KPIONE2_JUNE_OPERATIONAL_WEEKS_WITHOUT_RECORDS",
                "detail": kp["operational_weeks_missing_records"],
            })

    if blockers:
        status = "BLOCKED_FOR_UX_VISIBLE"
    elif warnings:
        status = "PARTIAL_FOR_UX_VISIBLE"
    else:
        status = "COMPLETE_FOR_UX_VISIBLE_GATE"

    return {
        "status": status,
        "blockers": blockers,
        "warnings": warnings,
    }


def write_markdown(report: dict[str, Any], md_path: Path) -> None:
    v = report["verdict"]

    lines = []
    lines.append(f"# {PHASE_ID}")
    lines.append("")
    lines.append(f"- generated_at: `{report['meta']['generated_at_local']}`")
    lines.append(f"- base: `{report['input_inventory']['base']}`")
    lines.append(f"- verdict: **{v['status']}**")
    lines.append("")
    lines.append("## Guardrails")
    for k, val in report["guardrails"].items():
        lines.append(f"- {k}: `{val}`")
    lines.append("")
    lines.append("## Calendario operativo junio 2026")
    lines.append("| semana | inicio | fin | días del mes |")
    lines.append("|---|---:|---:|---:|")
    for w in report["calendar"]["operational_weeks"]:
        lines.append(f"| {w['label']} | {w['week_start']} | {w['week_end']} | {w['days_in_month']} |")
    lines.append("")
    lines.append("## Inventario de entrada")
    inv = report["input_inventory"]
    lines.append(f"- ruta_files: `{len(inv['ruta_files'])}`")
    for p in inv["ruta_files"]:
        lines.append(f"  - `{p}`")
    lines.append(f"- photo_files: `{len(inv['photo_files'])}`")
    for p in inv["photo_files"]:
        lines.append(f"  - `{p}`")
    lines.append(f"- cumplimiento_file: `{inv['cumplimiento_file']}`")
    lines.append("")
    lines.append("## Perfiles")
    lines.append("| fuente | archivo | hoja | filas_data | columnas | missing_critical | fechas |")
    lines.append("|---|---|---|---:|---:|---|---|")

    for p in report["profiles"]:
        dp = p.get("date_profiles", {})
        dates = "; ".join(
            f"{k}:{v.get('min_date')}..{v.get('max_date')} ({v.get('parsed_count')})"
            for k, v in dp.items()
        ) or "-"

        lines.append(
            f"| {p.get('source_type')} | {p.get('file', {}).get('name')} | {p.get('sheet', '-')} | "
            f"{p.get('data_rows', '-')} | {p.get('max_column', '-')} | "
            f"{', '.join(p.get('missing_critical_fields', [])) or '-'} | {dates} |"
        )

    lines.append("")
    lines.append("## Cobertura junio")
    lines.append("| fuente | días con registros | semanas con registros | semanas faltantes |")
    lines.append("|---|---:|---|---|")

    for source, c in report["coverage"].items():
        lines.append(
            f"| {source} | {c['month_days_with_records']} | "
            f"{', '.join(c['operational_weeks_with_records']) or '-'} | "
            f"{', '.join(c['operational_weeks_missing_records']) or '-'} |"
        )

    lines.append("")
    lines.append("## Blockers")
    if v["blockers"]:
        for b in v["blockers"]:
            lines.append(f"- `{b['id']}`: {json.dumps(b, ensure_ascii=False)}")
    else:
        lines.append("- Sin blockers.")
    lines.append("")
    lines.append("## Warnings")
    if v["warnings"]:
        for w in v["warnings"]:
            lines.append(f"- `{w['id']}`: {json.dumps(w, ensure_ascii=False)}")
    else:
        lines.append("- Sin warnings.")
    lines.append("")
    lines.append("## Cierre")
    lines.append("Reporte no-apply. Solo evidencia local. No mueve archivos, no toca Supabase, no aplica SQL y no modifica loaders productivos.")
    lines.append("")

    md_path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--base", default=".")
    parser.add_argument("--out-dir", default=None)
    parser.add_argument("--max-rows", type=int, default=None)
    parser.add_argument("--soft-exit", action="store_true")
    args = parser.parse_args()

    base = Path(args.base).resolve()
    out_dir = Path(args.out_dir).resolve() if args.out_dir else base / "evidence" / PHASE_ID
    out_dir.mkdir(parents=True, exist_ok=True)

    weeks = operational_weeks_june_2026()
    inv = find_inputs(base)

    profiles = []

    for ruta in inv["ruta_files"]:
        profiles.append(
            profile_excel(
                Path(ruta),
                "ruta_rutero",
                ["RUTA RUTERO"],
                CRITICALS["ruta_rutero"],
                args.max_rows,
            )
        )

    for photo in inv["photo_files"]:
        profiles.append(
            profile_excel(
                Path(photo),
                "photo_admin",
                ["Fotos", "FOTOS"],
                CRITICALS["photo_admin"],
                args.max_rows,
            )
        )

    if inv["cumplimiento_file"]:
        cumplimiento = Path(inv["cumplimiento_file"])

        profiles.append(
            profile_excel(
                cumplimiento,
                "kpione2",
                ["DB (KPIONE2.0)"],
                CRITICALS["kpione2"],
                args.max_rows,
            )
        )

        profiles.append(
            profile_excel(
                cumplimiento,
                "power_app",
                ["DB (POWER_APP)"],
                CRITICALS["power_app"],
                args.max_rows,
            )
        )

        profiles.append(
            profile_excel(
                cumplimiento,
                "ruta_reference_tabla1",
                ["Tabla1"],
                CRITICALS["ruta_rutero"],
                args.max_rows,
            )
        )

    report = {
        "meta": {
            "phase_id": PHASE_ID,
            "generated_at_local": datetime.now().astimezone().isoformat(timespec="seconds"),
            "python": sys.version.split()[0],
            "platform": sys.platform,
        },
        "guardrails": {
            "supabase_write": False,
            "sql_apply": False,
            "ddl": False,
            "productive_loaders": False,
            "refresh_productivo": False,
            "contracts": False,
            "data_movement": False,
            "git_add_all": False,
            "stash_apply": False,
            "only_evidence_files_written": True,
        },
        "calendar": {
            "rule": "Semana operativa lunes-domingo; asignación al mes con >=4 días.",
            "month": "2026-06",
            "operational_weeks": weeks,
            "excluded_week_example": {
                "week_start": "2026-06-29",
                "week_end": "2026-07-05",
                "assigned_month": "2026-07",
            },
        },
        "input_inventory": inv,
        "profiles": profiles,
    }

    report["coverage"] = coverage(profiles, weeks)
    report["verdict"] = verdict(report)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = out_dir / f"{PHASE_ID}_{stamp}.json"
    md_path = out_dir / f"{PHASE_ID}_{stamp}.md"

    json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    write_markdown(report, md_path)

    print(json.dumps({
        "phase_id": PHASE_ID,
        "status": report["verdict"]["status"],
        "blockers": len(report["verdict"]["blockers"]),
        "warnings": len(report["verdict"]["warnings"]),
        "json": str(json_path),
        "markdown": str(md_path),
    }, ensure_ascii=False, indent=2))

    if args.soft_exit:
        return 0

    return 2 if report["verdict"]["status"].startswith("BLOCKED") else 0


if __name__ == "__main__":
    raise SystemExit(main())
