from __future__ import annotations

import hashlib
import json
import re
import subprocess
import unicodedata
import uuid
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Callable, Iterable
from urllib.parse import parse_qs, urlparse

from openpyxl import load_workbook


RUNNER_VERSION = "017_ROUTE_B_V1"
SOURCE_SHEET = "Fotos"
FILE_PATTERN = "photo-excel-admin_*.xlsx"
LOCAL_DB_ENV = "DB_URL_CODEX_LOCAL"
PRODUCTIVE_DB_ENV = "DB_URL_KPIONE_ROUTE_B_PRODUCTIVE"
PRODUCTIVE_CONFIRM_TOKEN = "KPIONE_ROUTE_B_018_APPLY"
PRODUCTIVE_ROLLBACK_CONFIRM_TOKEN = "KPIONE_ROUTE_B_018_ROLLBACK"
PLANNED_PRODUCTIVE_ROLE = "stock_zero_kpione_route_b_load"
EXPECTED_PRODUCTIVE_DATABASE = "postgres"
EXPECTED_PRODUCTIVE_PROJECT_REF = "xheyrgfagpoigpgakilu"
EXPECTED_PRODUCTIVE_HOSTNAME = "db.xheyrgfagpoigpgakilu.supabase.co"
PRODUCTIVE_PREPARATION_STATUS = "TECHNICAL_BOUNDARY_READY_ROLE_PROVISIONING_PENDING"
PRODUCTIVE_EXECUTION_STATUS = "READY_FOR_PRODUCTIVE_EXECUTION"
PRODUCTIVE_PREPARATION_BLOCKERS = [
    "PRODUCTIVE_ROLE_NOT_PROVISIONED_OR_VERIFIED",
    "READ_ONLY_PRECHECK_NOT_AUTHORIZED_OR_EXECUTED",
]
BANNED_PRODUCTIVE_ROLES = {
    "postgres", "stock_zero_codex_ro", "anon", "authenticated", "service_role",
}
ADVISORY_LOCK_KEY = 5426926728611921713  # Stable signed bigint for KPIONE_ROUTE_B_V1.
LIFECYCLE = (
    "DISCOVERED", "VALIDATING", "VALIDATED", "STAGING", "STAGED", "ACTIVE",
    "QUARANTINED", "SUPERSEDED", "ROLLED_BACK", "FAILED",
)

COLUMN_ALIASES = {
    "event_id": ("id",), "sp_item_id": ("sp item id",), "holding": ("holding",),
    "subcadena": ("subcadena",), "cod_rt": ("codigo local",),
    "cliente_norm": ("marca",), "local_nombre": ("local",),
    "direccion": ("direccion",), "reponedor": ("reponedor",), "fecha": ("fecha",),
    "fecha_subida": ("fecha de subida",), "hora": ("hora",),
    "tipo_de_tarea": ("tipo de tarea",),
    "photo_count": ("n fotos", "foto no/total", "foto n/total", "foto n o/total"),
    "comentarios": ("comentarios",), "link_foto": ("link foto",),
}
REQUIRED = tuple(k for k in COLUMN_ALIASES if k != "fecha_subida")
EVENT_STABLE = (
    "event_id", "sp_item_id", "holding", "subcadena", "cod_rt", "cliente_norm",
    "local_nombre", "direccion", "reponedor", "fecha", "comentarios",
)
PHOTO_FIELDS = ("photo_count", "link_foto", "hora", "tipo_de_tarea", "fecha_subida")


class RouteBError(RuntimeError):
    def __init__(self, identifier: str, *, connection_attempted: bool = False,
                 writes_attempted: bool = False, committed: bool = False,
                 report: dict[str, Any] | None = None) -> None:
        super().__init__(identifier)
        self.connection_attempted = connection_attempted
        self.writes_attempted = writes_attempted
        self.committed = committed
        self.report = report


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    text = unicodedata.normalize("NFC", str(value)).strip()
    return re.sub(r"\s+", " ", text)


def identity_key(value: Any) -> str:
    text = unicodedata.normalize("NFKD", clean_text(value))
    return "".join(c for c in text if not unicodedata.combining(c)).upper()


def normalize_column(value: Any) -> str:
    text = identity_key(value).lower().replace("º", "o").replace("°", "o")
    return re.sub(r"\s+", " ", text)


def normalize_numeric_string(value: Any) -> str:
    text = clean_text(value)
    if re.fullmatch(r"[+-]?\d+(?:\.0+)?", text):
        return str(int(float(text)))
    return text


def parse_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean_text(value)
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    raise RouteBError(f"invalid_date:{text}")


def stable_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def discover_files(input_dir: Path) -> list[Path]:
    if not input_dir.is_dir():
        raise RouteBError(f"input_dir_not_found:{input_dir}")
    candidates = sorted(input_dir.glob(FILE_PATTERN), key=lambda p: p.name.casefold())
    result: list[Path] = []
    resolved: set[Path] = set()
    for path in candidates:
        if path.name.startswith("~$"):
            continue
        real = path.resolve()
        if real in resolved:
            raise RouteBError(f"duplicate_resolved_path:{path}")
        resolved.add(real)
        result.append(path)
    if not result:
        raise RouteBError("no_route_b_files_discovered")
    return result


def assert_local_target(env_name: str, dsn: str | None) -> str:
    if env_name != LOCAL_DB_ENV:
        raise RouteBError(f"unsafe_db_env:{env_name}")
    if not dsn:
        raise RouteBError("missing_local_db_url")
    parsed = urlparse(dsn)
    host = (parsed.hostname or "").lower()
    if parsed.scheme not in {"postgres", "postgresql"}:
        raise RouteBError("unsupported_db_scheme")
    if "supabase" in dsn.lower() or host not in {"localhost", "127.0.0.1", "::1"}:
        raise RouteBError(f"non_local_db_target:{host or 'missing_host'}")
    params = parse_qs(parsed.query)
    if params.get("sslmode", [""])[0].lower() in {"require", "verify-ca", "verify-full"}:
        raise RouteBError("remote_ssl_metadata_rejected")
    return "LOCAL_POSTGRESQL_LOOPBACK"


def assert_postgresql_identifier(value: str) -> str:
    if not value or not re.fullmatch(r"[a-z_][a-z0-9_]{0,62}", value):
        raise RouteBError("invalid_postgresql_identifier")
    if value in BANNED_PRODUCTIVE_ROLES:
        raise RouteBError("banned_productive_role")
    return value


def load_approved_productive_plan(path: Path) -> dict[str, Any]:
    plan = json.loads(path.read_text(encoding="utf-8"))
    if plan.get("document_type") != "kpione_route_b_productive_apply_plan":
        raise RouteBError("invalid_productive_plan_type")
    if not isinstance(plan.get("productive_apply_authorized"), bool):
        raise RouteBError("productive_apply_authorization_flag_required")
    if not isinstance(plan.get("productive_rollback_authorized"), bool):
        raise RouteBError("productive_rollback_authorization_flag_required")
    return plan


def validate_productive_role_contract(plan: dict[str, Any]) -> str:
    target = plan.get("target", {})
    activation = plan.get("activation_gate", {})
    role = target.get("planned_productive_role")
    assert_postgresql_identifier(str(role or ""))
    if role != PLANNED_PRODUCTIVE_ROLE:
        raise RouteBError("planned_productive_role_mismatch")
    role_gate_state = (
        target.get("productive_role_status"),
        target.get("allowed_productive_roles"),
        activation.get("productive_role_registered"),
        activation.get("gate_open"),
    )
    readonly_precheck = plan.get("readonly_precheck")
    preparation_state = (
        plan.get("status") == PRODUCTIVE_PREPARATION_STATUS
        and plan.get("remaining_blockers") == PRODUCTIVE_PREPARATION_BLOCKERS
        and readonly_precheck == {
            "status": "NOT_AUTHORIZED_OR_EXECUTED",
            "evidence_sha256": None,
        }
        and role_gate_state == ("PLANNED_NOT_PROVISIONED", [], False, False)
        and plan.get("productive_apply_authorized") is False
        and plan.get("productive_rollback_authorized") is False
    )
    evidence_sha256 = (
        readonly_precheck.get("evidence_sha256")
        if isinstance(readonly_precheck, dict) else None
    )
    execution_state = (
        plan.get("status") == PRODUCTIVE_EXECUTION_STATUS
        and plan.get("remaining_blockers") == []
        and isinstance(readonly_precheck, dict)
        and readonly_precheck.get("status") == "PASSED"
        and isinstance(evidence_sha256, str)
        and re.fullmatch(r"[0-9a-f]{64}", evidence_sha256) is not None
        and role_gate_state == (
            "PROVISIONED_AND_VERIFIED", [PLANNED_PRODUCTIVE_ROLE], True, True,
        )
        and isinstance(plan.get("productive_apply_authorized"), bool)
        and isinstance(plan.get("productive_rollback_authorized"), bool)
    )
    if not preparation_state and not execution_state:
        raise RouteBError("productive_execution_state_contract_mismatch")
    return role


def validate_registered_productive_target(plan: dict[str, Any]) -> None:
    target = plan.get("target", {})
    expected = {
        "expected_supabase_project_ref": EXPECTED_PRODUCTIVE_PROJECT_REF,
        "expected_hostname": EXPECTED_PRODUCTIVE_HOSTNAME,
        "expected_database": EXPECTED_PRODUCTIVE_DATABASE,
    }
    for key, value in expected.items():
        if target.get(key) != value:
            raise RouteBError(f"productive_target_{key}_mismatch")
    if target.get("allowed_readonly_roles") != ["stock_zero_codex_ro"]:
        raise RouteBError("productive_readonly_role_contract_mismatch")


def validate_productive_dsn_target(dsn: str, plan: dict[str, Any]) -> dict[str, str]:
    parsed = urlparse(dsn)
    if parsed.scheme not in {"postgres", "postgresql"}:
        raise RouteBError("unsupported_productive_db_scheme")
    username = parsed.username or ""
    host = (parsed.hostname or "").lower()
    database = (parsed.path or "").lstrip("/")
    target = plan["target"]
    if username != target["planned_productive_role"]:
        raise RouteBError("productive_dsn_role_mismatch")
    if host != target["expected_hostname"]:
        raise RouteBError("productive_dsn_hostname_mismatch")
    if target["expected_supabase_project_ref"] not in host or "supabase" not in host:
        raise RouteBError("productive_dsn_project_ref_mismatch")
    if database != target["expected_database"]:
        raise RouteBError("productive_dsn_database_mismatch")
    sslmodes = parse_qs(parsed.query, keep_blank_values=True).get("sslmode", [])
    if sslmodes != ["require"]:
        raise RouteBError("productive_sslmode_require_required")
    return {
        "hostname": host,
        "database": database,
        "username": username,
        "project_ref": target["expected_supabase_project_ref"],
        "sslmode": sslmodes[0],
    }


def current_git_head(root: Path) -> str:
    return subprocess.check_output(
        ["git", "rev-parse", "HEAD"], cwd=root, text=True, stderr=subprocess.DEVNULL
    ).strip()


def _run_git(root: Path, *args: str, text: bool = True) -> subprocess.CompletedProcess[Any]:
    return subprocess.run(
        ["git", *args], cwd=root, capture_output=True, text=text, check=False,
    )


def validate_productive_git_guard(plan_path: Path, expected_git_ref: str,
                                  root: Path) -> dict[str, str]:
    if not expected_git_ref or not re.fullmatch(r"[0-9a-f]{40}", expected_git_ref):
        raise RouteBError("expected_plan_git_ref_required")
    head = current_git_head(root)
    if head != expected_git_ref:
        raise RouteBError("repository_head_mismatch")

    unstaged = _run_git(root, "diff", "--quiet")
    if unstaged.returncode not in {0, 1}:
        raise RouteBError("repository_worktree_status_unavailable")
    untracked = _run_git(root, "ls-files", "--others", "--exclude-standard")
    if untracked.returncode != 0:
        raise RouteBError("repository_worktree_status_unavailable")
    if unstaged.returncode == 1 or untracked.stdout.strip():
        raise RouteBError("repository_worktree_not_clean")
    staged = _run_git(root, "diff", "--cached", "--quiet")
    if staged.returncode not in {0, 1}:
        raise RouteBError("repository_index_status_unavailable")
    if staged.returncode == 1:
        raise RouteBError("repository_index_not_clean")

    try:
        relative_plan = plan_path.resolve().relative_to(root.resolve()).as_posix()
    except ValueError as exc:
        raise RouteBError("approved_plan_outside_repository") from exc
    tracked = _run_git(root, "cat-file", "-e", f"HEAD:{relative_plan}")
    if tracked.returncode != 0:
        raise RouteBError("approved_plan_not_tracked_at_head")
    head_blob = _run_git(root, "show", f"HEAD:{relative_plan}", text=False)
    if head_blob.returncode != 0:
        raise RouteBError("approved_plan_blob_unavailable")
    working_bytes = plan_path.read_bytes()
    if working_bytes != head_blob.stdout:
        raise RouteBError("approved_plan_worktree_blob_mismatch")
    return {
        "approved_git_sha": head,
        "plan_path": relative_plan,
        "plan_sha256": hashlib.sha256(working_bytes).hexdigest(),
    }


def sha256_lf_normalized(path: Path) -> str:
    return hashlib.sha256(path.read_bytes().replace(b"\r\n", b"\n")).hexdigest()


def _date_range(start: str, end: str) -> list[str]:
    current, final = date.fromisoformat(start), date.fromisoformat(end)
    values: list[str] = []
    while current <= final:
        values.append(current.isoformat())
        current += timedelta(days=1)
    return values


def _summarize_approved_workbooks(workbooks: list[WorkbookPlan]) -> dict[str, Any]:
    rows = [row for workbook in workbooks for row in workbook.rows]
    classified = classify_global_photo_duplicates(workbooks)
    canonical = [row for row in classified if row["duplicate_classification"] == "UNIQUE"]
    event_stability: dict[str, tuple[str, str]] = {}
    conflicts: set[str] = set()
    for row in rows:
        value = (row["sp_item_id"], row["event_stable_hash"])
        if row["event_id"] in event_stability and event_stability[row["event_id"]] != value:
            conflicts.add(row["event_id"])
        event_stability.setdefault(row["event_id"], value)
    dates = sorted({row["fecha"] for row in rows})
    presence = {(row["fecha"], row["location_key"], row["cliente_norm"]) for row in canonical}
    semantic_hashes = sorted(semantic_content_hash(workbook) for workbook in workbooks)
    semantic_plan_hash = sha256_text(stable_json({
        "runner_version": RUNNER_VERSION,
        "approved_semantic_hashes": semantic_hashes,
        "grain": "immutable_event_photo_staging_row",
    }))
    return {
        "approved_file_count": len(workbooks),
        "expected_source_rows": len(rows),
        "expected_distinct_events": len({row["event_id"] for row in canonical}),
        "expected_duplicate_photo_rows": sum(
            row["duplicate_classification"] != "UNIQUE" for row in classified
        ),
        "expected_exact_duplicate_rows": sum(
            row["duplicate_classification"] == "EXACT_DUPLICATE" for row in classified
        ),
        "expected_cross_file_duplicate_rows": sum(
            row["duplicate_classification"] == "CROSS_FILE_DUPLICATE" for row in classified
        ),
        "expected_event_conflicts": len(conflicts),
        "expected_day_presence_rows": len(presence),
        "coverage_start": dates[0],
        "coverage_end": dates[-1],
        "distinct_dates": len(dates),
        "missing_dates": sorted(set(_date_range(dates[0], dates[-1])) - set(dates)),
        "semantic_plan_hash": semantic_plan_hash,
        "_classified_rows": classified,
    }


def validate_productive_local_artifacts(plan: dict[str, Any], root: Path,
                                        *, validate_sql: bool = True) -> dict[str, Any]:
    if validate_sql:
        sql_path = root / plan["physical_contract"]["sql_file"]
        if not sql_path.is_file():
            raise RouteBError("sql_contract_missing")
        if sha256_lf_normalized(sql_path) != plan["physical_contract"]["sql_sha256"]:
            raise RouteBError("sql_sha256_mismatch")

    package = plan["source_package"]
    input_dir = root / plan["input_directory"]["repository_relative_value"]
    approved = {item["filename"]: item for item in package["approved_files"]}
    excluded = {item["filename"]: item for item in package["excluded_files"]}
    inventory = {item["filename"]: item for item in package["observed_directory_inventory"]}
    if set(approved) & set(excluded):
        raise RouteBError("approved_excluded_filename_overlap")
    if {item["sha256"] for item in approved.values()} & {
        item["sha256"] for item in excluded.values()
    }:
        raise RouteBError("approved_excluded_hash_overlap")
    actual = {
        path.name: path
        for path in sorted(input_dir.glob(plan["input_directory"]["file_pattern"]))
        if not path.name.startswith("~$")
    }
    unknown = sorted(set(actual) - set(inventory))
    if unknown:
        raise RouteBError("unknown_matching_files:" + ",".join(unknown))
    missing_approved = sorted(set(approved) - set(actual))
    if missing_approved:
        raise RouteBError("missing_approved_files:" + ",".join(missing_approved))
    missing_excluded = sorted(set(excluded) - set(actual))
    if missing_excluded:
        raise RouteBError("missing_excluded_files:" + ",".join(missing_excluded))
    if set(inventory) != set(approved) | set(excluded):
        raise RouteBError("manifest_inventory_membership_mismatch")

    for filename, path in actual.items():
        role = "APPROVED" if filename in approved else "EXCLUDED"
        if inventory[filename].get("manifest_role") != role:
            raise RouteBError("manifest_inventory_role_mismatch:" + filename)
        entry = approved.get(filename) or excluded[filename]
        if path.stat().st_size != entry.get("file_size"):
            raise RouteBError("observed_file_size_mismatch:" + filename)
        if sha256_file(path) != entry["sha256"] or entry["sha256"] != inventory[filename]["sha256"]:
            raise RouteBError("observed_file_hash_mismatch:" + filename)

    truncated = [item for item in excluded.values()
                 if item.get("classification") == "NEGATIVE_TEST_FIXTURE_TRUNCATED"]
    semantic_duplicates = [item for item in excluded.values()
                           if item.get("classification") == "DUPLICATE_SEMANTIC_EXPORT"]
    if len(truncated) != 1 or truncated[0]["filename"] in approved:
        raise RouteBError("truncated_fixture_exclusion_contract_mismatch")
    if len(semantic_duplicates) != 1:
        raise RouteBError("semantic_duplicate_exclusion_contract_mismatch")

    workbooks: list[WorkbookPlan] = []
    files: list[dict[str, Any]] = []
    for filename, entry in approved.items():
        workbook = inspect_workbook(actual[filename])
        if workbook.source_file_sha256 != entry["sha256"]:
            raise RouteBError("approved_raw_hash_mismatch:" + filename)
        if semantic_content_hash(workbook) != entry["semantic_content_hash"]:
            raise RouteBError("approved_semantic_hash_mismatch:" + filename)
        for field, observed in (
            ("file_size", workbook.file_size),
            ("row_count", len(workbook.rows)),
            ("distinct_event_count", workbook.event_count),
            ("day_presence_count", workbook.day_presence_count),
            ("duplicate_photo_rows", workbook.duplicate_rows),
            ("coverage_start", workbook.coverage_start),
            ("coverage_end", workbook.coverage_end),
        ):
            if entry.get(field) != observed:
                raise RouteBError(f"approved_metric_mismatch:{filename}:{field}")
        workbooks.append(workbook)
        files.append({
            "source_file_name": workbook.source_file_name,
            "source_file_sha256": workbook.source_file_sha256,
            "source_sheet": workbook.source_sheet,
            "file_size": workbook.file_size,
            "coverage_start": workbook.coverage_start,
            "coverage_end": workbook.coverage_end,
            "row_count": len(workbook.rows),
            "event_count": workbook.event_count,
            "classification": "APPROVED_MANIFEST_SOURCE",
        })

    duplicate_entry = semantic_duplicates[0]
    duplicate_workbook = inspect_workbook(actual[duplicate_entry["filename"]])
    duplicate_semantic_hash = semantic_content_hash(duplicate_workbook)
    canonical_name = duplicate_entry.get("semantic_equivalent_approved_file")
    if (duplicate_semantic_hash != duplicate_entry.get("semantic_content_hash")
            or canonical_name not in approved
            or approved[canonical_name]["semantic_content_hash"] != duplicate_semantic_hash):
        raise RouteBError("semantic_duplicate_exclusion_contract_mismatch")

    summary = _summarize_approved_workbooks(workbooks)
    classified = summary.pop("_classified_rows")
    summary.update({
        "approved_source_bytes": sum(actual[name].stat().st_size for name in approved),
        "excluded_source_bytes": sum(actual[name].stat().st_size for name in excluded),
        "observed_directory_bytes": sum(path.stat().st_size for path in actual.values()),
        "excluded_file_count": len(excluded),
        "observed_file_count": len(actual),
    })
    expected_summary = {
        "approved_file_count": package["approved_file_count"],
        "expected_source_rows": package["expected_source_rows"],
        "expected_distinct_events": package["expected_distinct_events"],
        "expected_duplicate_photo_rows": package["expected_duplicate_photo_rows"],
        "expected_exact_duplicate_rows": package["expected_exact_duplicate_rows"],
        "expected_cross_file_duplicate_rows": package["expected_cross_file_duplicate_rows"],
        "expected_event_conflicts": package["expected_event_conflicts"],
        "expected_day_presence_rows": package["expected_day_presence_rows"],
        "coverage_start": package["expected_coverage"]["start"],
        "coverage_end": package["expected_coverage"]["end"],
        "distinct_dates": package["expected_coverage"]["distinct_dates"],
        "missing_dates": package["expected_coverage"]["missing_dates"],
        "semantic_plan_hash": package["semantic_plan_hash"],
        **package["source_byte_metrics"],
    }
    if summary != expected_summary:
        raise RouteBError("approved_aggregate_mismatch")
    return {
        "runner_version": RUNNER_VERSION,
        "semantic_plan_hash": package["semantic_plan_hash"],
        "apply_authorized": False,
        "db_target_classification": "PRODUCTIVE_TARGET_NOT_CONNECTED",
        "coverage_start": package["expected_coverage"]["start"],
        "coverage_end": package["expected_coverage"]["end"],
        "discovered_file_count": len(workbooks),
        "already_active_file_count": 0,
        "renamed_no_op_count": 0,
        "new_file_count": len(workbooks),
        "files_selected_for_staging": len(workbooks),
        "files_skipped_as_no_op": 0,
        "source_rows": package["expected_source_rows"],
        "duplicate_rows": package["expected_duplicate_photo_rows"],
        "distinct_events": package["expected_distinct_events"],
        "event_conflicts": package["expected_event_conflicts"],
        "day_presence_count": package["expected_day_presence_rows"],
        "expected_inserts": package["expected_source_rows"],
        "expected_no_ops": 0,
        "expected_quarantines": 0,
        "expected_supersession_requirement": False,
        "files": files,
        "_workbooks": workbooks,
        "_classified_rows": classified,
        "_validation_summary": summary,
    }


def build_approved_plan_from_manifest(plan: dict[str, Any], root: Path) -> dict[str, Any]:
    return validate_productive_local_artifacts(plan, root)


def productive_blocked_report(plan: dict[str, Any], mode: str) -> dict[str, Any]:
    return {
        "verdict": "BLOCKED",
        "mode": mode,
        "status": plan.get("status"),
        "productive_apply_authorized": False,
        "planned_productive_role": plan.get("target", {}).get("planned_productive_role"),
        "productive_role_status": plan.get("target", {}).get("productive_role_status"),
        "allowed_productive_roles": plan.get("target", {}).get("allowed_productive_roles", []),
        "remaining_blockers": plan.get("remaining_blockers", []),
        "connection_attempted": False,
        "writes_attempted": False,
        "committed": False,
        "dsn_read": False,
    }


def require_productive_gate_open(plan: dict[str, Any], mode: str) -> None:
    validate_productive_role_contract(plan)
    validate_registered_productive_target(plan)
    if plan.get("status") != PRODUCTIVE_EXECUTION_STATUS:
        raise RouteBError(f"{mode}_gate_closed")
    if mode == "apply_productive" and plan.get("productive_apply_authorized") is not True:
        raise RouteBError("productive_apply_not_authorized")
    if mode == "rollback_productive" and plan.get("productive_rollback_authorized") is not True:
        raise RouteBError("productive_rollback_not_authorized")


def _productive_connect(dsn: str, connect_fn: Callable[[str], Any] | None) -> Any:
    if connect_fn is None:
        import psycopg
        connect_fn = psycopg.connect
    return connect_fn(dsn)


def _validate_productive_session(cursor: Any, plan: dict[str, Any],
                                 *, require_readonly: bool) -> dict[str, str]:
    cursor.execute(
        "SELECT current_user,session_user,current_database(),"
        "current_setting('transaction_read_only')"
    )
    current_user, session_user, database, readonly = cursor.fetchone()
    expected_role = plan["target"]["planned_productive_role"]
    if current_user != expected_role or session_user != expected_role:
        raise RouteBError("productive_session_role_mismatch")
    if database != plan["target"]["expected_database"]:
        raise RouteBError("productive_session_database_mismatch")
    expected_readonly = "on" if require_readonly else "off"
    if readonly != expected_readonly:
        raise RouteBError(
            "productive_postcheck_not_read_only" if require_readonly
            else "productive_apply_transaction_not_read_write"
        )
    return {
        "current_user": str(current_user),
        "session_user": str(session_user),
        "database": str(database),
        "transaction_read_only": str(readonly),
    }


def _assert_route_b_object_signatures(cursor: Any, plan: dict[str, Any],
                                      *, allow_empty: bool) -> None:
    expected = plan["physical_contract"]["object_signatures"]
    names = sorted(expected)
    cursor.execute(
        "SELECT n.nspname||'.'||c.relname,c.relkind FROM pg_class c "
        "JOIN pg_namespace n ON n.oid=c.relnamespace "
        "WHERE n.nspname||'.'||c.relname = ANY(%s) ORDER BY 1",
        (names,),
    )
    actual_kinds = {name: kind for name, kind in cursor.fetchall()}
    if not actual_kinds and allow_empty:
        return
    if actual_kinds != {name: spec["relation_kind"] for name, spec in expected.items()}:
        raise RouteBError("route_b_object_set_or_kind_mismatch")
    cursor.execute(
        "SELECT table_schema||'.'||table_name,column_name FROM information_schema.columns "
        "WHERE table_schema||'.'||table_name = ANY(%s) "
        "ORDER BY table_schema,table_name,ordinal_position",
        (names,),
    )
    actual_columns: dict[str, list[str]] = {}
    for relation, column in cursor.fetchall():
        actual_columns.setdefault(relation, []).append(column)
    if actual_columns != {name: spec["columns"] for name, spec in expected.items()}:
        raise RouteBError("route_b_object_column_signature_mismatch")


def _expected_productive_counts(plan: dict[str, Any]) -> dict[str, int]:
    package = plan["source_package"]
    return {
        "staged_rows": package["expected_source_rows"],
        "unique_rows": package["expected_source_rows"] - package["expected_duplicate_photo_rows"],
        "exact_duplicate_rows": package["expected_exact_duplicate_rows"],
        "cross_file_duplicate_rows": package["expected_cross_file_duplicate_rows"],
        "total_duplicate_rows": package["expected_duplicate_photo_rows"],
        "events": package["expected_distinct_events"],
        "day_presence": package["expected_day_presence_rows"],
        "conflicts": package["expected_event_conflicts"],
        "identities_without_exactly_one_unique": 0,
        "files": package["approved_file_count"],
    }


def _observed_productive_counts(cursor: Any, batch_id: str) -> dict[str, int]:
    cursor.execute(
        "SELECT count(*),"
        "count(*) FILTER (WHERE duplicate_classification='UNIQUE'),"
        "count(*) FILTER (WHERE duplicate_classification='EXACT_DUPLICATE'),"
        "count(*) FILTER (WHERE duplicate_classification='CROSS_FILE_DUPLICATE'),"
        "count(*) FILTER (WHERE duplicate_classification IN ('EXACT_DUPLICATE','CROSS_FILE_DUPLICATE')),"
        "count(DISTINCT event_id) FILTER (WHERE duplicate_classification='UNIQUE'),"
        "count(DISTINCT (event_date,location_key,cliente_norm)) "
        "  FILTER (WHERE duplicate_classification='UNIQUE'),"
        "count(*) FILTER (WHERE conflict_classification<>'NONE') "
        "FROM cg_raw.kpione_raw_event_photo_staging_v1 WHERE batch_id=%s",
        (batch_id,),
    )
    values = cursor.fetchone()
    cursor.execute(
        "SELECT count(*) FROM (SELECT event_id,photo_row_hash FROM "
        "cg_raw.kpione_raw_event_photo_staging_v1 WHERE batch_id=%s "
        "GROUP BY event_id,photo_row_hash "
        "HAVING count(*) FILTER (WHERE duplicate_classification='UNIQUE')<>1) q",
        (batch_id,),
    )
    identities_without_unique = cursor.fetchone()[0]
    cursor.execute(
        "SELECT count(*) FROM cg_raw.kpione_raw_ingest_batch_file_v1 WHERE batch_id=%s",
        (batch_id,),
    )
    files = cursor.fetchone()[0]
    keys = (
        "staged_rows", "unique_rows", "exact_duplicate_rows",
        "cross_file_duplicate_rows", "total_duplicate_rows", "events",
        "day_presence", "conflicts",
    )
    result = {key: int(value) for key, value in zip(keys, values)}
    result["identities_without_exactly_one_unique"] = int(identities_without_unique)
    result["files"] = int(files)
    return result


def _target_fingerprint(plan: dict[str, Any]) -> str:
    target = plan["target"]
    return sha256_text(stable_json({
        "project_ref": target["expected_supabase_project_ref"],
        "hostname": target["expected_hostname"],
        "database": target["expected_database"],
        "role": target["planned_productive_role"],
    }))


def _write_productive_evidence(path: Path, report: dict[str, Any]) -> None:
    try:
        path.write_text(
            json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
            encoding="utf-8",
        )
    except OSError:
        raise RouteBError(
            "productive_evidence_write_failed",
            connection_attempted=True,
            writes_attempted=True,
            committed=True,
            report=report,
        ) from None


def _run_productive_postcheck(plan: dict[str, Any], dsn: str, batch_id: str,
                              operation: str, connect_fn: Callable[[str], Any] | None,
                              predecessor_batch_id: str | None = None) -> tuple[dict[str, Any], dict[str, str]]:
    try:
        connection = _productive_connect(dsn, connect_fn)
    except Exception:
        raise RouteBError(
            "productive_postcheck_connection_failed",
            connection_attempted=True,
            writes_attempted=True,
            committed=True,
        ) from None
    try:
        connection.autocommit = False
        with connection.cursor() as cursor:
            cursor.execute("BEGIN READ ONLY")
            cursor.execute("SET LOCAL statement_timeout = '5min'")
            cursor.execute("SET LOCAL lock_timeout = '5s'")
            session = _validate_productive_session(cursor, plan, require_readonly=True)
            declared: list[dict[str, Any]] = []
            declared_failures: list[int] = []
            for index, query in enumerate(plan["postcheck_queries"]):
                cursor.execute(query)
                rows = cursor.fetchall()
                result: dict[str, Any] = {"index": index, "row_count": len(rows)}
                if (
                    len(rows) == 1
                    and len(rows[0]) == 1
                    and isinstance(rows[0][0], bool)
                ):
                    boolean_result = rows[0][0]
                    result.update({
                        "result_type": "boolean",
                        "boolean_result": boolean_result,
                    })
                    if not boolean_result:
                        declared_failures.append(index)
                declared.append(result)
            cursor.execute(
                "SELECT status,supersedes_batch_id::text FROM "
                "cg_raw.kpione_raw_ingest_batch_v1 WHERE batch_id=%s",
                (batch_id,),
            )
            batch_row = cursor.fetchone()
            if not batch_row:
                raise RouteBError("productive_postcheck_batch_missing")
            status, observed_predecessor = batch_row
            observed = _observed_productive_counts(cursor, batch_id)
            cursor.execute(
                "SELECT count(*) FROM cg_raw.kpione_raw_ingest_batch_v1 WHERE status='ACTIVE'"
            )
            active_batches = int(cursor.fetchone()[0])
            cursor.execute(
                "SELECT count(*) FROM cg_raw.kpione_raw_ingest_batch_v1 a "
                "JOIN cg_raw.kpione_raw_ingest_batch_v1 b ON a.batch_id<b.batch_id "
                "AND a.status='ACTIVE' AND b.status='ACTIVE' "
                "AND daterange(a.coverage_start,a.coverage_end,'[]') "
                "&& daterange(b.coverage_start,b.coverage_end,'[]')"
            )
            overlapping_active_batches = int(cursor.fetchone()[0])
            cursor.execute("SELECT to_regclass('cg_raw.kpione2_raw')::text")
            legacy_object = cursor.fetchone()[0]
            expected = _expected_productive_counts(plan)
            if operation == "APPLY":
                passed = (
                    status == "ACTIVE"
                    and observed == expected
                    and active_batches == 1
                    and overlapping_active_batches == 0
                    and declared_failures == []
                    and legacy_object == "cg_raw.kpione2_raw"
                )
            else:
                predecessor_restored = True
                if predecessor_batch_id:
                    cursor.execute(
                        "SELECT count(*) FROM cg_raw.kpione_raw_ingest_batch_v1 "
                        "WHERE batch_id=%s AND status='ACTIVE'",
                        (predecessor_batch_id,),
                    )
                    predecessor_restored = cursor.fetchone()[0] == 1
                passed = (
                    status == "ROLLED_BACK"
                    and observed["staged_rows"] == expected["staged_rows"]
                    and observed_predecessor == predecessor_batch_id
                    and predecessor_restored
                    and active_batches <= 1
                    and overlapping_active_batches == 0
                    and declared_failures == []
                    and legacy_object == "cg_raw.kpione2_raw"
                )
            details = {
                "declared_postchecks": declared,
                "declared_failures": declared_failures,
                "batch_status": status,
                "active_batch_count": active_batches,
                "overlapping_active_batch_count": overlapping_active_batches,
                "legacy_object": legacy_object,
                "expected_counts": expected,
                "observed_counts": observed,
                "verdict": "PASS" if passed else "REJECTED",
            }
        connection.rollback()
        return details, session
    except RouteBError:
        connection.rollback()
        raise
    except Exception:
        connection.rollback()
        raise RouteBError(
            "productive_postcheck_failed",
            connection_attempted=True,
            writes_attempted=True,
            committed=True,
        ) from None
    finally:
        connection.close()


def _productive_evidence(plan: dict[str, Any], approved_plan: dict[str, Any],
                         git_guard: dict[str, str], operation: str,
                         execution_id: str, batch_id: str,
                         session: dict[str, str], postcheck: dict[str, Any],
                         *, predecessor_batch_id: str | None = None) -> dict[str, Any]:
    passed = postcheck["verdict"] == "PASS"
    return {
        "document_type": "kpione_route_b_productive_execution_evidence_v1",
        "operation": operation,
        "execution_uuid": execution_id,
        "timestamp_utc": datetime.now(timezone.utc).isoformat(),
        "approved_git_sha": git_guard["approved_git_sha"],
        "plan_path": git_guard["plan_path"],
        "plan_sha256": git_guard["plan_sha256"],
        "sql_sha256": plan["physical_contract"]["sql_sha256"],
        "target_fingerprint": _target_fingerprint(plan),
        "current_user": session["current_user"],
        "session_user": session["session_user"],
        "source_hashes": sorted(
            item["source_file_sha256"] for item in approved_plan["files"]
        ),
        "batch_id": batch_id,
        "predecessor_batch_id": predecessor_batch_id,
        "expected_counts": postcheck["expected_counts"],
        "observed_counts": postcheck["observed_counts"],
        "transaction_outcome": "COMMITTED",
        "committed": True,
        "postcheck_verdict": postcheck["verdict"],
        "postcheck": postcheck,
        "rollback_readiness": (
            "EXPLICIT_LOGICAL_ROLLBACK_AVAILABLE" if operation == "APPLY"
            else "LOGICAL_ROLLBACK_COMPLETED"
        ),
        "downstream_use_allowed": operation == "APPLY" and passed,
        "connection_attempted": True,
        "writes_attempted": True,
    }


def run_productive_apply(plan: dict[str, Any], approved_plan: dict[str, Any],
                         dsn: str, postcheck_report_json: Path,
                         *, git_guard: dict[str, str], root: Path,
                         connect_fn: Callable[[str], Any] | None = None) -> dict[str, Any]:
    validate_productive_dsn_target(dsn, plan)
    execution_id = str(uuid.uuid4())
    batch_id = str(uuid.uuid4())
    connection: Any | None = None
    writes_attempted = False
    committed = False
    session: dict[str, str] = {}
    try:
        connection = _productive_connect(dsn, connect_fn)
        connection.autocommit = False
        with connection.cursor() as cursor:
            session = _validate_productive_session(cursor, plan, require_readonly=False)
            cursor.execute("SET LOCAL statement_timeout = '15min'")
            cursor.execute("SET LOCAL lock_timeout = '10s'")
            cursor.execute("SET LOCAL idle_in_transaction_session_timeout = '5min'")
            cursor.execute("SELECT pg_advisory_xact_lock(%s)", (ADVISORY_LOCK_KEY,))
            _assert_route_b_object_signatures(cursor, plan, allow_empty=False)
            writes_attempted = True
            cursor.execute(
                "SELECT batch_id::text FROM cg_raw.kpione_raw_ingest_batch_v1 "
                "WHERE status='ACTIVE' FOR UPDATE"
            )
            if cursor.fetchall():
                raise RouteBError("unexpected_active_route_b_batch")
            package = plan["source_package"]
            cursor.execute(
                "INSERT INTO cg_raw.kpione_raw_ingest_batch_v1("
                "batch_id,runner_execution_id,semantic_plan_hash,status,coverage_start,coverage_end,"
                "file_count,row_count,event_count,validated_at) "
                "VALUES(%s,%s,%s,'STAGING',%s,%s,%s,%s,%s,clock_timestamp())",
                (
                    batch_id, execution_id, package["semantic_plan_hash"],
                    package["expected_coverage"]["start"],
                    package["expected_coverage"]["end"],
                    package["approved_file_count"], package["expected_source_rows"],
                    package["expected_distinct_events"],
                ),
            )
            file_rows = [
                (
                    batch_id, workbook.source_file_sha256, workbook.source_file_name,
                    workbook.source_sheet, workbook.file_size, workbook.coverage_start,
                    workbook.coverage_end, len(workbook.rows), workbook.event_count,
                )
                for workbook in approved_plan["_workbooks"]
            ]
            cursor.executemany(
                "INSERT INTO cg_raw.kpione_raw_ingest_batch_file_v1("
                "batch_id,source_file_sha256,source_file_name,source_sheet,file_size,"
                "coverage_start,coverage_end,row_count,event_count,validation_status) "
                "VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,'VALIDATED')",
                file_rows,
            )
            staging_rows = [
                (
                    batch_id, row["source_file_sha256"], row["source_sheet"],
                    row["source_row_number"], row["source_row_identity"], row["event_id"],
                    row["sp_item_id"], stable_json(row), row["photo_row_hash"],
                    row["event_stable_hash"], row["fecha"], row["location_key"],
                    row["cliente_norm"], row["duplicate_classification"],
                    row["conflict_classification"],
                )
                for row in approved_plan["_classified_rows"]
            ]
            cursor.executemany(
                "INSERT INTO cg_raw.kpione_raw_event_photo_staging_v1("
                "batch_id,source_file_sha256,source_sheet,source_row_number,source_row_identity,"
                "event_id,sp_item_id,source_payload,photo_row_hash,event_stable_hash,event_date,"
                "location_key,cliente_norm,duplicate_classification,conflict_classification) "
                "VALUES(%s,%s,%s,%s,%s,%s,%s,%s::jsonb,%s,%s,%s,%s,%s,%s,%s)",
                staging_rows,
            )
            observed = _observed_productive_counts(cursor, batch_id)
            if observed != _expected_productive_counts(plan):
                raise RouteBError("productive_staging_count_mismatch")
            cursor.execute(
                "UPDATE cg_raw.kpione_raw_ingest_batch_v1 "
                "SET status='ACTIVE',activated_at=clock_timestamp() "
                "WHERE batch_id=%s AND status='STAGING'",
                (batch_id,),
            )
            if cursor.rowcount != 1:
                raise RouteBError("productive_activation_failed")
        connection.commit()
        committed = True
    except RouteBError as exc:
        if connection is not None and not committed:
            connection.rollback()
        raise RouteBError(
            str(exc), connection_attempted=connection is not None,
            writes_attempted=writes_attempted, committed=committed,
        ) from None
    except Exception:
        if connection is not None and not committed:
            connection.rollback()
        raise RouteBError(
            "productive_apply_transaction_failed",
            connection_attempted=connection is not None,
            writes_attempted=writes_attempted,
            committed=committed,
        ) from None
    finally:
        if connection is not None:
            connection.close()

    try:
        postcheck, postcheck_session = _run_productive_postcheck(
            plan, dsn, batch_id, "APPLY", connect_fn,
        )
    except RouteBError as exc:
        raise RouteBError(
            str(exc), connection_attempted=True, writes_attempted=True,
            committed=True, report=exc.report,
        ) from None
    evidence = _productive_evidence(
        plan, approved_plan, git_guard, "APPLY", execution_id, batch_id,
        postcheck_session or session, postcheck,
    )
    _write_productive_evidence(postcheck_report_json, evidence)
    if postcheck["verdict"] != "PASS":
        raise RouteBError(
            "POSTCHECK_REJECTED_REQUIRES_EXPLICIT_ROLLBACK_AUTHORIZATION",
            connection_attempted=True, writes_attempted=True, committed=True,
            report=evidence,
        )
    return evidence


def run_productive_rollback(plan: dict[str, Any], approved_plan: dict[str, Any],
                            rollback_batch_id: str, dsn: str,
                            postcheck_report_json: Path, *, git_guard: dict[str, str],
                            connect_fn: Callable[[str], Any] | None = None) -> dict[str, Any]:
    try:
        rollback_batch_id = str(uuid.UUID(rollback_batch_id))
    except (ValueError, AttributeError):
        raise RouteBError("rollback_batch_uuid_required") from None
    validate_productive_dsn_target(dsn, plan)
    execution_id = str(uuid.uuid4())
    connection: Any | None = None
    writes_attempted = False
    committed = False
    predecessor: str | None = None
    session: dict[str, str] = {}
    try:
        connection = _productive_connect(dsn, connect_fn)
        connection.autocommit = False
        with connection.cursor() as cursor:
            session = _validate_productive_session(cursor, plan, require_readonly=False)
            cursor.execute("SET LOCAL statement_timeout = '5min'")
            cursor.execute("SET LOCAL lock_timeout = '10s'")
            cursor.execute("SET LOCAL idle_in_transaction_session_timeout = '5min'")
            cursor.execute("SELECT pg_advisory_xact_lock(%s)", (ADVISORY_LOCK_KEY,))
            _assert_route_b_object_signatures(cursor, plan, allow_empty=False)
            cursor.execute(
                "SELECT status,supersedes_batch_id::text FROM "
                "cg_raw.kpione_raw_ingest_batch_v1 WHERE batch_id=%s FOR UPDATE",
                (rollback_batch_id,),
            )
            rows = cursor.fetchall()
            if len(rows) != 1 or rows[0][0] != "ACTIVE":
                raise RouteBError("rollback_batch_not_active")
            predecessor = rows[0][1]
            cursor.execute(
                "SELECT count(*) FROM cg_raw.kpione_raw_event_photo_staging_v1 "
                "WHERE batch_id=%s",
                (rollback_batch_id,),
            )
            staging_before = int(cursor.fetchone()[0])
            if staging_before != plan["source_package"]["expected_source_rows"]:
                raise RouteBError("rollback_staging_immutability_precheck_failed")
            if predecessor:
                cursor.execute(
                    "SELECT status FROM cg_raw.kpione_raw_ingest_batch_v1 "
                    "WHERE batch_id=%s FOR UPDATE",
                    (predecessor,),
                )
                predecessor_rows = cursor.fetchall()
                if len(predecessor_rows) != 1 or predecessor_rows[0][0] != "SUPERSEDED":
                    raise RouteBError("rollback_predecessor_not_restorable")
            writes_attempted = True
            cursor.execute(
                "UPDATE cg_raw.kpione_raw_ingest_batch_v1 "
                "SET status='ROLLED_BACK',rolled_back_at=clock_timestamp() "
                "WHERE batch_id=%s AND status='ACTIVE'",
                (rollback_batch_id,),
            )
            if cursor.rowcount != 1:
                raise RouteBError("rollback_batch_transition_failed")
            if predecessor:
                cursor.execute(
                    "UPDATE cg_raw.kpione_raw_ingest_batch_v1 "
                    "SET status='ACTIVE',activated_at=clock_timestamp() "
                    "WHERE batch_id=%s AND status='SUPERSEDED'",
                    (predecessor,),
                )
                if cursor.rowcount != 1:
                    raise RouteBError("rollback_predecessor_not_restorable")
            cursor.execute(
                "SELECT count(*) FROM cg_raw.kpione_raw_event_photo_staging_v1 "
                "WHERE batch_id=%s",
                (rollback_batch_id,),
            )
            if int(cursor.fetchone()[0]) != staging_before:
                raise RouteBError("rollback_mutated_staging")
        connection.commit()
        committed = True
    except RouteBError as exc:
        if connection is not None and not committed:
            connection.rollback()
        raise RouteBError(
            str(exc), connection_attempted=connection is not None,
            writes_attempted=writes_attempted, committed=committed,
        ) from None
    except Exception:
        if connection is not None and not committed:
            connection.rollback()
        raise RouteBError(
            "productive_rollback_transaction_failed",
            connection_attempted=connection is not None,
            writes_attempted=writes_attempted,
            committed=committed,
        ) from None
    finally:
        if connection is not None:
            connection.close()

    try:
        postcheck, postcheck_session = _run_productive_postcheck(
            plan, dsn, rollback_batch_id, "ROLLBACK", connect_fn, predecessor,
        )
    except RouteBError as exc:
        raise RouteBError(
            str(exc), connection_attempted=True, writes_attempted=True,
            committed=True, report=exc.report,
        ) from None
    evidence = _productive_evidence(
        plan, approved_plan, git_guard, "ROLLBACK", execution_id,
        rollback_batch_id, postcheck_session or session, postcheck,
        predecessor_batch_id=predecessor,
    )
    _write_productive_evidence(postcheck_report_json, evidence)
    if postcheck["verdict"] != "PASS":
        raise RouteBError(
            "POSTCHECK_REJECTED_REQUIRES_EXPLICIT_ROLLBACK_AUTHORIZATION",
            connection_attempted=True, writes_attempted=True, committed=True,
            report=evidence,
        )
    return evidence


def _resolved_columns(headers: Iterable[Any]) -> tuple[dict[str, str], list[str], list[str]]:
    actual = {normalize_column(v): clean_text(v) for v in headers if clean_text(v)}
    resolved: dict[str, str] = {}
    for key, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if normalize_column(alias) in actual:
                resolved[key] = actual[normalize_column(alias)]
                break
    missing = [key for key in REQUIRED if key not in resolved]
    used = set(resolved.values())
    optional = sorted(v for v in actual.values() if v not in used)
    return resolved, missing, optional


@dataclass(frozen=True)
class WorkbookPlan:
    source_path: Path
    source_file_sha256: str
    source_file_name: str
    source_sheet: str
    file_size: int
    coverage_start: str
    coverage_end: str
    rows: tuple[dict[str, Any], ...]
    event_count: int
    day_presence_count: int
    duplicate_rows: int
    optional_columns: tuple[str, ...]


def inspect_workbook(path: Path) -> WorkbookPlan:
    content_hash = sha256_file(path)  # Identity is fixed before workbook parsing.
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise RouteBError(f"workbook_open_failed:{path.name}:{exc}") from exc
    try:
        if SOURCE_SHEET not in workbook.sheetnames:
            raise RouteBError(f"missing_sheet:{SOURCE_SHEET}")
        sheet = workbook[SOURCE_SHEET]
        values = sheet.iter_rows(values_only=True)
        headers = next(values, None)
        if not headers:
            raise RouteBError("missing_header")
        resolved, missing, optional = _resolved_columns(headers)
        if missing:
            raise RouteBError("missing_required_columns:" + ",".join(missing))
        positions = {clean_text(value): idx for idx, value in enumerate(headers)}
        rows: list[dict[str, Any]] = []
        event_stability: dict[str, tuple[str, str]] = {}
        event_dates: dict[str, set[str]] = {}
        seen_photo: set[tuple[str, str]] = set()
        duplicates = 0
        for source_row_number, values_row in enumerate(values, start=2):
            raw = {key: values_row[positions[column]] for key, column in resolved.items()}
            event_id = normalize_numeric_string(raw["event_id"])
            sp_item_id = normalize_numeric_string(raw["sp_item_id"])
            if not event_id or not sp_item_id:
                raise RouteBError(f"missing_identity:row_{source_row_number}")
            fecha = parse_date(raw["fecha"])
            normalized = {key: clean_text(value) for key, value in raw.items()}
            normalized.update({
                "event_id": event_id, "sp_item_id": sp_item_id, "fecha": fecha,
                "cod_rt_norm": identity_key(raw["cod_rt"]),
                "cliente_norm": identity_key(raw["cliente_norm"]),
                "local_nombre_norm": identity_key(raw["local_nombre"]),
            })
            normalized["location_key"] = normalized["cod_rt_norm"] or normalized["local_nombre_norm"]
            stable_payload = {key: normalized.get(key, "") for key in EVENT_STABLE}
            photo_payload = {key: normalized.get(key, "") for key in PHOTO_FIELDS}
            event_stable_hash = sha256_text(stable_json(stable_payload))
            photo_row_hash = sha256_text(stable_json(photo_payload))
            previous = event_stability.setdefault(event_id, (sp_item_id, event_stable_hash))
            if previous != (sp_item_id, event_stable_hash):
                raise RouteBError(f"event_stability_conflict:{event_id}")
            event_dates.setdefault(event_id, set()).add(fecha)
            if len(event_dates[event_id]) > 1:
                raise RouteBError(f"event_multi_date_conflict:{event_id}")
            photo_identity = (event_id, photo_row_hash)
            duplicate = photo_identity in seen_photo
            duplicates += int(duplicate)
            seen_photo.add(photo_identity)
            source_row_identity = sha256_text(stable_json([content_hash, SOURCE_SHEET, source_row_number]))
            rows.append({
                **normalized,
                "source_row_number": source_row_number,
                "source_row_identity": source_row_identity,
                "photo_row_hash": photo_row_hash,
                "event_stable_hash": event_stable_hash,
                "duplicate_classification": "EXACT_DUPLICATE" if duplicate else "UNIQUE",
                "conflict_classification": "NONE",
            })
        if not rows:
            raise RouteBError("empty_source_file")
        dates = sorted({row["fecha"] for row in rows})
        presence = {(r["fecha"], r["location_key"], r["cliente_norm"]) for r in rows}
        return WorkbookPlan(path, content_hash, path.name, SOURCE_SHEET, path.stat().st_size,
                            dates[0], dates[-1], tuple(rows), len(event_stability), len(presence),
                            duplicates, tuple(optional))
    finally:
        workbook.close()


def build_plan(input_dir: Path) -> dict[str, Any]:
    discovered_plans = [inspect_workbook(path) for path in discover_files(input_dir)]
    unique: dict[str, WorkbookPlan] = {}
    for workbook in discovered_plans:
        unique.setdefault(workbook.source_file_sha256, workbook)
    plans = list(unique.values())
    all_rows = [row for plan in plans for row in plan.rows]
    event_stability: dict[str, tuple[str, str]] = {}
    for row in all_rows:
        value = (row["sp_item_id"], row["event_stable_hash"])
        if event_stability.setdefault(row["event_id"], value) != value:
            raise RouteBError(f"cross_file_event_stability_conflict:{row['event_id']}")
    coverage = sorted({row["fecha"] for row in all_rows})
    presence = {(r["fecha"], r["location_key"], r["cliente_norm"]) for r in all_rows}
    source_versions = sorted(p.source_file_sha256 for p in plans)
    semantic = {"runner_version": RUNNER_VERSION, "source_versions": source_versions,
                "source_sheet": SOURCE_SHEET, "grain": "immutable_event_photo_staging_row"}
    return {
        "runner_version": RUNNER_VERSION,
        "semantic_plan_hash": sha256_text(stable_json(semantic)),
        "apply_authorized": False,
        "db_target_classification": "NOT_EVALUATED_DRY_RUN",
        "coverage_start": coverage[0], "coverage_end": coverage[-1],
        "discovered_file_count": len(discovered_plans),
        "already_active_file_count": 0,
        "renamed_no_op_count": 0,
        "new_file_count": len(plans),
        "files_selected_for_staging": len(plans),
        "files_skipped_as_no_op": len(discovered_plans) - len(plans),
        "source_rows": len(all_rows), "duplicate_rows": sum(p.duplicate_rows for p in plans),
        "distinct_events": len(event_stability), "event_conflicts": 0,
        "day_presence_count": len(presence), "expected_inserts": len(all_rows),
        "expected_no_ops": 0, "expected_quarantines": 0,
        "expected_supersession_requirement": False,
        "files": [{
            "source_file_name": p.source_file_name, "source_file_sha256": p.source_file_sha256,
            "source_sheet": p.source_sheet, "file_size": p.file_size,
            "coverage_start": p.coverage_start, "coverage_end": p.coverage_end,
            "row_count": len(p.rows), "event_count": p.event_count,
            "optional_columns": list(p.optional_columns),
            "classification": (
                "NEW_SOURCE_VERSION" if unique[p.source_file_sha256] is p
                else "DUPLICATE_DISCOVERED_SOURCE_VERSION"
            ),
        } for p in discovered_plans],
        "_workbooks": plans,
        "_discovered_workbooks": discovered_plans,
    }


def public_plan(plan: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in plan.items() if not key.startswith("_")}


def semantic_content_records(workbook: WorkbookPlan) -> list[tuple[str, ...]]:
    return sorted(
        (
            row["event_id"],
            row["sp_item_id"],
            row["event_stable_hash"],
            row["photo_row_hash"],
            row["fecha"],
            row["location_key"],
            row["cliente_norm"],
        )
        for row in workbook.rows
    )


def semantic_content_hash(workbook: WorkbookPlan) -> str:
    return sha256_text(stable_json(semantic_content_records(workbook)))


def classify_global_photo_duplicates(workbooks: Iterable[WorkbookPlan]) -> list[dict[str, Any]]:
    occurrences: list[tuple[str, str, int, dict[str, Any]]] = []
    for workbook in workbooks:
        for row in workbook.rows:
            occurrences.append((
                workbook.source_file_sha256,
                workbook.source_sheet,
                int(row["source_row_number"]),
                row,
            ))
    occurrences.sort(key=lambda item: (item[0], item[1], item[2]))
    canonical_source: dict[tuple[str, str], str] = {}
    classified: list[dict[str, Any]] = []
    for source_hash, source_sheet, source_row_number, row in occurrences:
        identity = (row["event_id"], row["photo_row_hash"])
        if identity not in canonical_source:
            classification = "UNIQUE"
            canonical_source[identity] = source_hash
        elif canonical_source[identity] == source_hash:
            classification = "EXACT_DUPLICATE"
        else:
            classification = "CROSS_FILE_DUPLICATE"
        classified.append({
            **row,
            "source_file_sha256": source_hash,
            "source_sheet": source_sheet,
            "source_row_number": source_row_number,
            "duplicate_classification": classification,
        })
    return classified


def _semantic_plan_hash(workbooks: list[WorkbookPlan]) -> str:
    semantic = {
        "runner_version": RUNNER_VERSION,
        "source_versions": sorted(p.source_file_sha256 for p in workbooks),
        "source_sheet": SOURCE_SHEET,
        "grain": "immutable_event_photo_staging_row",
    }
    return sha256_text(stable_json(semantic))


def _apply_report(plan: dict[str, Any], classifications: list[dict[str, Any]],
                  selected: list[WorkbookPlan]) -> dict[str, Any]:
    already_active = [item for item in classifications if item["classification"].startswith("ALREADY_ACTIVE")]
    renamed = [item for item in classifications if item["classification"] == "ALREADY_ACTIVE_DIFFERENT_NAME"]
    selected_rows = sum(len(workbook.rows) for workbook in selected)
    report = public_plan(plan)
    report.update({
        "files": classifications,
        "already_active_file_count": len(already_active),
        "renamed_no_op_count": len(renamed),
        "new_file_count": len(selected),
        "files_selected_for_staging": len(selected),
        "files_skipped_as_no_op": len(classifications) - len(selected),
        "expected_inserts": selected_rows,
        "expected_no_ops": len(classifications) - len(selected),
        "semantic_plan_hash": _semantic_plan_hash(selected) if selected else None,
    })
    return report


def apply_local(plan: dict[str, Any], dsn: str, ddl_path: Path,
                supersede_batch_id: str | None = None) -> dict[str, Any]:
    import psycopg
    target = assert_local_target(LOCAL_DB_ENV, dsn)
    execution_id = str(uuid.uuid4())
    batch_id = str(uuid.uuid4())
    discovered: list[WorkbookPlan] = plan["_discovered_workbooks"]
    with psycopg.connect(dsn) as connection:
        with connection.transaction(), connection.cursor() as cursor:
            cursor.execute("SELECT pg_advisory_xact_lock(%s)", (ADVISORY_LOCK_KEY,))
            cursor.execute(ddl_path.read_text(encoding="utf-8"))
            hashes = sorted({p.source_file_sha256 for p in discovered})
            cursor.execute("SELECT b.batch_id::text, f.source_file_sha256, f.source_file_name FROM cg_raw.kpione_raw_ingest_batch_v1 b JOIN cg_raw.kpione_raw_ingest_batch_file_v1 f USING(batch_id) WHERE b.status='ACTIVE' AND f.source_file_sha256 = ANY(%s)", (hashes,))
            active_versions: dict[str, dict[str, set[str]]] = {}
            for active_batch_id, source_hash, source_name in cursor.fetchall():
                item = active_versions.setdefault(source_hash, {"names": set(), "batch_ids": set()})
                item["names"].add(source_name)
                item["batch_ids"].add(active_batch_id)
            classifications: list[dict[str, Any]] = []
            selected_by_hash: dict[str, WorkbookPlan] = {}
            for workbook in discovered:
                active = active_versions.get(workbook.source_file_sha256)
                if active:
                    classification = (
                        "ALREADY_ACTIVE_SAME_NAME"
                        if workbook.source_file_name in active["names"]
                        else "ALREADY_ACTIVE_DIFFERENT_NAME"
                    )
                elif workbook.source_file_sha256 in selected_by_hash:
                    classification = "DUPLICATE_DISCOVERED_SOURCE_VERSION"
                else:
                    classification = "NEW_SOURCE_VERSION"
                    selected_by_hash[workbook.source_file_sha256] = workbook
                classifications.append({
                    "source_file_name": workbook.source_file_name,
                    "source_file_sha256": workbook.source_file_sha256,
                    "coverage_start": workbook.coverage_start,
                    "coverage_end": workbook.coverage_end,
                    "row_count": len(workbook.rows),
                    "classification": classification,
                })
            workbooks = list(selected_by_hash.values())
            report = _apply_report(plan, classifications, workbooks)
            if not workbooks:
                outcome = (
                    "NO_OP_SAME_SOURCE_VERSION"
                    if report["renamed_no_op_count"]
                    else "NO_OP_ALREADY_REGISTERED"
                )
                batch_ids = sorted({batch_id for item in active_versions.values() for batch_id in item["batch_ids"]})
                return {**report, "outcome": outcome, "active_batch_ids": batch_ids,
                        "apply_authorized": True, "db_target_classification": target}
            coverage_start = min(p.coverage_start for p in workbooks)
            coverage_end = max(p.coverage_end for p in workbooks)
            cursor.execute("SELECT batch_id::text FROM cg_raw.kpione_raw_ingest_batch_v1 WHERE status='ACTIVE' AND daterange(coverage_start, coverage_end, '[]') && daterange(%s,%s,'[]')", (coverage_start, coverage_end))
            overlaps = [row[0] for row in cursor.fetchall()]
            if overlaps and not supersede_batch_id:
                for item in classifications:
                    if item["classification"] == "NEW_SOURCE_VERSION":
                        item["classification"] = "NEW_SOURCE_VERSION_REQUIRES_SUPERSESSION"
                report = _apply_report(plan, classifications, workbooks)
                report["expected_supersession_requirement"] = True
                return {**report, "outcome": "NEW_SOURCE_VERSION_PENDING_SUPERSESSION",
                        "active_batch_ids": overlaps, "apply_authorized": False,
                        "db_target_classification": target}
            if supersede_batch_id and (supersede_batch_id not in overlaps or len(overlaps) != 1):
                raise RouteBError("invalid_or_unrelated_supersession")
            selected_rows = [row for workbook in workbooks for row in workbook.rows]
            selected_events = {row["event_id"] for row in selected_rows}
            cursor.execute("INSERT INTO cg_raw.kpione_raw_ingest_batch_v1(batch_id,runner_execution_id,semantic_plan_hash,status,coverage_start,coverage_end,file_count,row_count,event_count,validated_at,supersedes_batch_id) VALUES(%s,%s,%s,'STAGING',%s,%s,%s,%s,%s,clock_timestamp(),%s)", (batch_id, execution_id, report["semantic_plan_hash"], coverage_start, coverage_end, len(workbooks), len(selected_rows), len(selected_events), supersede_batch_id))
            for workbook in workbooks:
                cursor.execute("INSERT INTO cg_raw.kpione_raw_ingest_batch_file_v1(batch_id,source_file_sha256,source_file_name,source_sheet,file_size,coverage_start,coverage_end,row_count,event_count,validation_status) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,'VALIDATED')", (batch_id, workbook.source_file_sha256, workbook.source_file_name, workbook.source_sheet, workbook.file_size, workbook.coverage_start, workbook.coverage_end, len(workbook.rows), workbook.event_count))
                for row in workbook.rows:
                    cursor.execute("INSERT INTO cg_raw.kpione_raw_event_photo_staging_v1(batch_id,source_file_sha256,source_sheet,source_row_number,source_row_identity,event_id,sp_item_id,source_payload,photo_row_hash,event_stable_hash,event_date,location_key,cliente_norm,duplicate_classification,conflict_classification) VALUES(%s,%s,%s,%s,%s,%s,%s,%s::jsonb,%s,%s,%s,%s,%s,%s,%s)", (batch_id, workbook.source_file_sha256, workbook.source_sheet, row["source_row_number"], row["source_row_identity"], row["event_id"], row["sp_item_id"], stable_json(row), row["photo_row_hash"], row["event_stable_hash"], row["fecha"], row["location_key"], row["cliente_norm"], row["duplicate_classification"], row["conflict_classification"]))
            if supersede_batch_id:
                cursor.execute("UPDATE cg_raw.kpione_raw_ingest_batch_v1 SET status='SUPERSEDED' WHERE batch_id=%s AND status='ACTIVE'", (supersede_batch_id,))
                if cursor.rowcount != 1:
                    raise RouteBError("supersession_predecessor_not_active")
            cursor.execute("UPDATE cg_raw.kpione_raw_ingest_batch_v1 SET status='ACTIVE',activated_at=clock_timestamp() WHERE batch_id=%s AND status='STAGING'", (batch_id,))
    return {**report, "outcome": "ACTIVE", "batch_id": batch_id, "apply_authorized": True,
            "db_target_classification": target}


def rollback_local(dsn: str, batch_id: str) -> dict[str, Any]:
    import psycopg
    target = assert_local_target(LOCAL_DB_ENV, dsn)
    with psycopg.connect(dsn) as connection:
        with connection.transaction(), connection.cursor() as cursor:
            cursor.execute("SELECT pg_advisory_xact_lock(%s)", (ADVISORY_LOCK_KEY,))
            cursor.execute("SELECT supersedes_batch_id::text FROM cg_raw.kpione_raw_ingest_batch_v1 WHERE batch_id=%s AND status='ACTIVE' FOR UPDATE", (batch_id,))
            row = cursor.fetchone()
            if not row:
                raise RouteBError("rollback_batch_not_active")
            predecessor = row[0]
            cursor.execute("UPDATE cg_raw.kpione_raw_ingest_batch_v1 SET status='ROLLED_BACK',rolled_back_at=clock_timestamp() WHERE batch_id=%s", (batch_id,))
            if predecessor:
                cursor.execute("UPDATE cg_raw.kpione_raw_ingest_batch_v1 SET status='ACTIVE',activated_at=clock_timestamp() WHERE batch_id=%s AND status='SUPERSEDED'", (predecessor,))
                if cursor.rowcount != 1:
                    raise RouteBError("rollback_predecessor_not_restorable")
            cursor.execute("SELECT count(*) FROM cg_core.kpione_event_normalized_v1")
            events = cursor.fetchone()[0]
            cursor.execute("SELECT count(*) FROM cg_core.kpione_day_presence_v1")
            presence = cursor.fetchone()[0]
    return {"outcome": "ROLLED_BACK", "restored_batch_id": predecessor, "event_count": events,
            "day_presence_count": presence, "db_target_classification": target}
