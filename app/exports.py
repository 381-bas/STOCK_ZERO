# app/exports.py
import io
import re
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, A3, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, LongTable, TableStyle, Paragraph, Spacer
from reportlab.lib.units import cm


EXPORT_COLS = [
    "Fecha stock",
    "MARCA",
    "Sku",
    "Descripción del Producto",
    "Stock",
    "VENTA 0",
    "NEGATIVO",
    "RIESGO DE QUIEBRE",
    "OTROS",
]

FOCUS_EXPORT_COLS = [
    "Fecha stock",
    "COD_RT",
    "LOCAL",
    "CLIENTE",
    "RUTERO",
    "REPONEDOR",
    "MARCA",
    "Sku",
    "Descripción del Producto",
    "Stock",
    "FOCO PRINCIPAL",
    "ACCIÓN SUGERIDA",
    "VENTA 0",
    "NEGATIVO",
    "RIESGO DE QUIEBRE",
    "OTROS",
]

INVENTORY_CLIENTE_EXPORT_COLS = [
    "Fecha stock",
    "COD_RT",
    "LOCAL",
    "CLIENTE",
    "GESTOR",
    "SUPERVISOR",
    "RUTERO",
    "REPONEDOR",
    "MODALIDAD",
    "MARCA",
    "Sku",
    "Descripción del Producto",
    "Stock",
    "Venta(+7)",
    "VENTA 0",
    "FOCO PRINCIPAL",
    "ACCIÓN SUGERIDA",
    "NEGATIVO",
    "RIESGO DE QUIEBRE",
    "OTROS",
]

CG_V2_DETAIL_EXPORT_COLS = [
    "SEMANA",
    "GESTOR",
    "RUTERO",
    "COD_RT",
    "LOCAL",
    "CLIENTE",
    "MODALIDAD",
    "EXIGIDAS SEM.",
    "LUN",
    "MAR",
    "MIE",
    "JUE",
    "VIE",
    "SAB",
    "DOM",
    "PENDIENTE",
    "ALERTA",
    "GESTION COMPARTIDA",
    "RUTA COMPARTIDA",
]


def _collapse_pipe_values(values) -> str:
    parts: list[str] = []
    seen: set[str] = set()
    for raw in values:
        text = str(raw or "").strip()
        if not text or text.lower() in {"nan", "none"}:
            continue
        for item in text.split("|"):
            token = item.strip()
            if not token:
                continue
            key = token.upper()
            if key in seen:
                continue
            seen.add(key)
            parts.append(token)
    return " | ".join(parts)


def _coalesce_duplicate_rr_columns(df_in: pd.DataFrame) -> pd.DataFrame:
    if df_in is None:
        return df_in

    df = df_in.copy()
    for base in ["GESTOR", "SUPERVISOR", "RUTERO", "REPONEDOR", "MODALIDAD"]:
        variants = [c for c in [base, f"{base}_x", f"{base}_y"] if c in df.columns]
        if len(variants) <= 1:
            continue

        cols = list(df.columns)
        insert_at = min(cols.index(c) for c in variants)
        merged = df[variants].apply(lambda row: _collapse_pipe_values(row.tolist()), axis=1)

        drop_cols = [c for c in variants if c != base]
        if base in df.columns:
            df[base] = merged
        else:
            df.insert(insert_at, base, merged)
        if drop_cols:
            df = df.drop(columns=drop_cols, errors="ignore")

    return df


def _clean_yes(v) -> str:
    return "SI" if str(v or "").strip().upper() == "SI" else ""


def _clean_otros(v) -> str:
    out = str(v or "").strip()
    return "" if out.upper() in {"", "NO", "N/A", "NA", "-"} else out


def _venta_0_flag(v) -> str:
    try:
        return "SI" if int(pd.to_numeric(v, errors="coerce") or 0) == 0 else ""
    except Exception:
        return ""


def build_export_df(df_ux: pd.DataFrame) -> pd.DataFrame:
    df = df_ux.copy()

    needed = [
        "fecha",
        "MARCA",
        "Sku",
        "Descripción del Producto",
        "Stock",
        "Venta(+7)",
        "NEGATIVO",
        "RIESGO DE QUIEBRE",
        "OTROS",
    ]
    for c in needed:
        if c not in df.columns:
            df[c] = ""

    df["Fecha stock"] = pd.to_datetime(df["fecha"], errors="coerce").dt.strftime("%Y-%m-%d").fillna("")
    df["Sku"] = df["Sku"].astype(str)
    df["Stock"] = pd.to_numeric(df["Stock"], errors="coerce").fillna(0).astype(int)
    df["Venta(+7)"] = pd.to_numeric(df["Venta(+7)"], errors="coerce").fillna(0).astype(int)

    df["VENTA 0"] = df["Venta(+7)"].apply(_venta_0_flag)
    df["NEGATIVO"] = df["NEGATIVO"].apply(_clean_yes)
    df["RIESGO DE QUIEBRE"] = df["RIESGO DE QUIEBRE"].apply(_clean_yes)
    df["OTROS"] = df["OTROS"].apply(_clean_otros)

    return df[EXPORT_COLS]


def _sorted_for_export(df_in: pd.DataFrame) -> pd.DataFrame:
    if df_in is None or df_in.empty:
        return df_in

    df = df_in.copy()

    if "MARCA" not in df.columns:
        df["MARCA"] = ""
    if "Sku" not in df.columns:
        df["Sku"] = ""
    if "Descripción del Producto" not in df.columns:
        df["Descripción del Producto"] = ""

    for c in ["MARCA", "Sku", "Descripción del Producto", "CLIENTE", "COD_RT", "LOCAL", "RUTERO", "REPONEDOR", "FOCO PRINCIPAL"]:
        if c in df.columns:
            df[c] = df[c].astype(str)

    sku_num = pd.to_numeric(df["Sku"], errors="coerce")
    df["_sku_is_text"] = sku_num.isna().astype(int)
    df["_sku_num"] = sku_num.fillna(0)

    if "FOCO PRINCIPAL" in df.columns and ({"CLIENTE", "COD_RT", "LOCAL"} & set(df.columns)):
        if "CLIENTE" not in df.columns:
            df["CLIENTE"] = ""
        if "COD_RT" not in df.columns:
            df["COD_RT"] = ""
        if "LOCAL" not in df.columns:
            df["LOCAL"] = ""
        if "RUTERO" not in df.columns:
            df["RUTERO"] = ""
        if "REPONEDOR" not in df.columns:
            df["REPONEDOR"] = ""
        sort_cols = [
            "CLIENTE", "COD_RT", "LOCAL", "RUTERO", "REPONEDOR",
            "MARCA", "FOCO PRINCIPAL", "_sku_is_text", "_sku_num", "Sku", "Descripción del Producto"
        ]
    elif "FOCO PRINCIPAL" in df.columns:
        sort_cols = ["MARCA", "FOCO PRINCIPAL", "_sku_is_text", "_sku_num", "Sku", "Descripción del Producto"]
    else:
        sort_cols = ["MARCA", "_sku_is_text", "_sku_num", "Sku", "Descripción del Producto"]

    df = df.sort_values(
        by=sort_cols,
        ascending=[True] * len(sort_cols),
        kind="mergesort",
    ).drop(columns=["_sku_is_text", "_sku_num"])

    return df


def _focus_principal(row) -> str:
    neg = _clean_yes(row.get("NEGATIVO", ""))
    quiebre = _clean_yes(row.get("RIESGO DE QUIEBRE", ""))
    venta0 = _venta_0_flag(row.get("Venta(+7)", 0))
    otros = _clean_otros(row.get("OTROS", ""))

    # prioridad semántica operativa
    if neg == "SI":
        return "NEGATIVO"
    if quiebre == "SI":
        return "QUIEBRE"
    if venta0 == "SI":
        return "VENTA 0"
    if otros:
        return "OTROS"
    return ""


def _accion_sugerida(foco_principal: str) -> str:
    if foco_principal == "NEGATIVO":
        return "Ajustar inventario y validar sala"
    if foco_principal == "QUIEBRE":
        return "Solicitar empuje o reposición"
    if foco_principal == "VENTA 0":
        return "Revisar exhibición, precio y rotación"
    if foco_principal == "OTROS":
        return "Revisar observación cliente"
    return ""


def _normalize_focos_export(foco) -> list[str]:
    valid = ["Venta 0", "Negativo", "Quiebres", "Otros"]

    if foco is None:
        raw = []
    elif isinstance(foco, str):
        raw = [x.strip() for x in foco.replace("|", ",").split(",") if x.strip()]
    else:
        raw = [str(x).strip() for x in foco if str(x).strip()]

    out = []
    for x in raw:
        if x in valid and x not in out:
            out.append(x)
    return out


def build_inventory_cliente_export_df(df_ux: pd.DataFrame) -> pd.DataFrame:
    df = df_ux.copy()

    needed = [
        "fecha",
        "COD_RT",
        "LOCAL",
        "CLIENTE",
        "GESTOR",
        "SUPERVISOR",
        "RUTERO",
        "REPONEDOR",
        "MODALIDAD",
        "MARCA",
        "Sku",
        "Descripción del Producto",
        "Stock",
        "Venta(+7)",
        "NEGATIVO",
        "RIESGO DE QUIEBRE",
        "OTROS",
    ]
    for c in needed:
        if c not in df.columns:
            df[c] = ""

    df["Fecha stock"] = pd.to_datetime(df["fecha"], errors="coerce").dt.strftime("%Y-%m-%d").fillna("")
    for c in [
        "COD_RT",
        "LOCAL",
        "CLIENTE",
        "GESTOR",
        "SUPERVISOR",
        "RUTERO",
        "REPONEDOR",
        "MODALIDAD",
        "MARCA",
    ]:
        df[c] = df[c].astype(str).replace({"nan": "", "None": ""}).fillna("")
    df["Sku"] = df["Sku"].astype(str)
    df["Stock"] = pd.to_numeric(df["Stock"], errors="coerce").fillna(0).astype(int)
    df["Venta(+7)"] = pd.to_numeric(df["Venta(+7)"], errors="coerce").fillna(0).astype(int)

    df["VENTA 0"] = df["Venta(+7)"].apply(_venta_0_flag)
    df["NEGATIVO"] = df["NEGATIVO"].apply(_clean_yes)
    df["RIESGO DE QUIEBRE"] = df["RIESGO DE QUIEBRE"].apply(_clean_yes)
    df["OTROS"] = df["OTROS"].apply(_clean_otros)
    df["FOCO PRINCIPAL"] = df.apply(_focus_principal, axis=1)
    df["ACCIÓN SUGERIDA"] = df["FOCO PRINCIPAL"].apply(_accion_sugerida)

    if df.empty:
        return pd.DataFrame(columns=INVENTORY_CLIENTE_EXPORT_COLS)

    df = _sorted_for_export(df)
    return df[INVENTORY_CLIENTE_EXPORT_COLS]


def build_focus_export_df(df_ux: pd.DataFrame, foco: str | list[str] = "Todo") -> pd.DataFrame:
    df = df_ux.copy()

    needed = [
        "fecha",
        "COD_RT",
        "LOCAL",
        "CLIENTE",
        "RUTERO",
        "REPONEDOR",
        "MARCA",
        "Sku",
        "Descripción del Producto",
        "Stock",
        "Venta(+7)",
        "NEGATIVO",
        "RIESGO DE QUIEBRE",
        "OTROS",
    ]
    for c in needed:
        if c not in df.columns:
            df[c] = ""

    df["Fecha stock"] = pd.to_datetime(df["fecha"], errors="coerce").dt.strftime("%Y-%m-%d").fillna("")
    for c in ["COD_RT", "LOCAL", "CLIENTE", "RUTERO", "REPONEDOR", "MARCA"]:
        df[c] = df[c].astype(str).replace({"nan": "", "None": ""}).fillna("")
    df["Sku"] = df["Sku"].astype(str)
    df["Stock"] = pd.to_numeric(df["Stock"], errors="coerce").fillna(0).astype(int)
    df["Venta(+7)"] = pd.to_numeric(df["Venta(+7)"], errors="coerce").fillna(0).astype(int)

    df["VENTA 0"] = df["Venta(+7)"].apply(_venta_0_flag)
    df["NEGATIVO"] = df["NEGATIVO"].apply(_clean_yes)
    df["RIESGO DE QUIEBRE"] = df["RIESGO DE QUIEBRE"].apply(_clean_yes)
    df["OTROS"] = df["OTROS"].apply(_clean_otros)
    df["FOCO PRINCIPAL"] = df.apply(_focus_principal, axis=1)
    df["ACCIÓN SUGERIDA"] = df["FOCO PRINCIPAL"].apply(_accion_sugerida)

    focos = _normalize_focos_export(foco)

    if focos:
        mask = pd.Series(False, index=df.index)

        if "Venta 0" in focos:
            mask = mask | (df["VENTA 0"] == "SI")
        if "Negativo" in focos:
            mask = mask | (df["NEGATIVO"] == "SI")
        if "Quiebres" in focos:
            mask = mask | (df["RIESGO DE QUIEBRE"] == "SI")
        if "Otros" in focos:
            mask = mask | (df["OTROS"] != "")

        df = df[mask].copy()
    else:
        df = df[df["FOCO PRINCIPAL"] != ""].copy()

    if df.empty:
        return pd.DataFrame(columns=FOCUS_EXPORT_COLS)

    df = _sorted_for_export(df)
    return df[FOCUS_EXPORT_COLS]


def _safe_file_token(value: object, *, fallback: str = "scope") -> str:
    token = str(value or "").strip()
    token = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", token)
    token = re.sub(r"\s+", "_", token)
    token = token.strip("._")
    return token or fallback


def _cg_v2_export_day_value(value: object) -> str:
    raw = str(value or "").strip().upper()
    if raw in {"1 ✓", "1 \u2713"}:
        return "1 \u2713"
    if raw == "REQ_OK":
        return "1 \u2713"
    if raw == "REQ":
        return "1"
    if raw in {"OK", "\u2713"}:
        return "\u2713"
    return "" if raw in {"", "NAN", "NONE"} else str(value)


def _cg_v2_prepare_detail_export_df(detail_df: pd.DataFrame | None) -> pd.DataFrame:
    if detail_df is None or detail_df.empty:
        return pd.DataFrame(columns=CG_V2_DETAIL_EXPORT_COLS)

    df = detail_df.copy()
    for col in CG_V2_DETAIL_EXPORT_COLS:
        if col not in df.columns:
            df[col] = ""
    if "SEMANA" in df.columns:
        df["SEMANA"] = pd.to_datetime(df["SEMANA"], errors="coerce").dt.strftime("%Y-%m-%d").fillna(df["SEMANA"].astype(str))
    for col in ["EXIGIDAS SEM.", "PENDIENTE"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
    for day_col in ["LUN", "MAR", "MIE", "JUE", "VIE", "SAB", "DOM"]:
        df[day_col] = df[day_col].apply(_cg_v2_export_day_value)
    for col in CG_V2_DETAIL_EXPORT_COLS:
        if col not in {"EXIGIDAS SEM.", "PENDIENTE"}:
            df[col] = df[col].fillna("")
    return df[CG_V2_DETAIL_EXPORT_COLS]


def _cg_v2_count_non_blank_unique(series: pd.Series) -> int:
    values = {
        str(value).strip()
        for value in series.tolist()
        if str(value or "").strip() and str(value).strip().lower() not in {"nan", "none"}
    }
    return len(values)


def _cg_v2_context_export_df(context: dict[str, Any]) -> pd.DataFrame:
    rows = [
        ("Semana operativa", context.get("Semana operativa") or ""),
        ("Gestor", context.get("Gestor") or "Todos"),
        ("Vista de analisis", context.get("Vista de analisis") or ""),
        ("Foco", context.get("Foco") or "scope"),
        ("Alerta", context.get("Alerta") or "Todas"),
        ("Fuente weekly", context.get("Fuente weekly") or ""),
        ("Generado en", context.get("Generado en") or ""),
    ]
    return pd.DataFrame(rows, columns=["CAMPO", "VALOR"])


def _cg_v2_summary_export_df(summary: dict[str, Any]) -> pd.DataFrame:
    visita_plan = int(summary.get("visita_plan") or 0)
    visita_realizada_cap = int(summary.get("visita_realizada_cap") or 0)
    pct_cumplimiento = round((visita_realizada_cap / visita_plan) * 100, 1) if visita_plan > 0 else 0.0
    rows = [
        ("Filas raw weekly", int(summary.get("total_rows") or 0)),
        ("Visitas exigidas semanales", visita_plan),
        ("Visitas válidas", visita_realizada_cap),
        ("% cumplimiento", pct_cumplimiento),
        ("Visitas pendientes", int(summary.get("visitas_pendientes") or 0)),
        ("Rutas cumplen", int(summary.get("cumple_rows") or 0)),
        ("Rutas incumplen", int(summary.get("incumple_rows") or 0)),
        ("Rutas no aplica", int(summary.get("no_aplica_rows") or 0)),
        ("Sobrecumplimiento", int(summary.get("sobre_cumplimiento") or 0)),
        ("Gestión compartida (rows)", int(summary.get("gestion_compartida_rows") or 0)),
    ]
    return pd.DataFrame(rows, columns=["METRICA", "VALOR"])


def _cg_v2_top_incumplimientos_por_cliente(detail_df: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "RANK",
        "CLIENTE",
        "LOCALES VISIBLES",
        "COD_RT VISIBLES",
        "EXIGIDAS SEM. TOTAL",
        "PENDIENTE TOTAL",
        "INCUMPLE ROWS",
    ]
    if detail_df is None or detail_df.empty:
        return pd.DataFrame(columns=columns)

    df = detail_df[detail_df["ALERTA"].astype(str).str.upper() == "INCUMPLE"].copy()
    if df.empty:
        return pd.DataFrame(columns=columns)

    grouped = (
        df.groupby("CLIENTE", dropna=False)
        .agg(
            **{
                "LOCALES VISIBLES": ("LOCAL", _cg_v2_count_non_blank_unique),
                "COD_RT VISIBLES": ("COD_RT", _cg_v2_count_non_blank_unique),
                "EXIGIDAS SEM. TOTAL": ("EXIGIDAS SEM.", "sum"),
                "PENDIENTE TOTAL": ("PENDIENTE", "sum"),
                "INCUMPLE ROWS": ("ALERTA", "count"),
            }
        )
        .reset_index()
    )
    grouped["CLIENTE"] = grouped["CLIENTE"].fillna("").astype(str)
    grouped = grouped.sort_values(
        by=["PENDIENTE TOTAL", "INCUMPLE ROWS", "CLIENTE"],
        ascending=[False, False, True],
        kind="mergesort",
    ).head(10)
    grouped.insert(0, "RANK", range(1, len(grouped) + 1))
    return grouped[columns]


def _cg_v2_top_locales_rutas_pendiente(detail_df: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "RANK",
        "COD_RT",
        "LOCAL",
        "CLIENTE",
        "GESTOR",
        "RUTERO",
        "RUTA COMPARTIDA",
        "EXIGIDAS SEM.",
        "PENDIENTE",
        "ALERTA",
    ]
    if detail_df is None or detail_df.empty:
        return pd.DataFrame(columns=columns)

    df = detail_df[detail_df["ALERTA"].astype(str).str.upper() != "NO APLICA"].copy()
    if df.empty:
        return pd.DataFrame(columns=columns)

    df = df.sort_values(
        by=["PENDIENTE", "EXIGIDAS SEM.", "CLIENTE", "LOCAL"],
        ascending=[False, False, True, True],
        kind="mergesort",
    ).head(10)
    df = df.loc[:, [col for col in columns if col != "RANK"]].copy()
    df.insert(0, "RANK", range(1, len(df) + 1))
    return df[columns]


def _cg_v2_get_or_create_sheet(writer: pd.ExcelWriter, sheet_name: str):
    safe_sheet = str(sheet_name or "DATA")[:31]
    workbook = writer.book
    if safe_sheet in writer.sheets:
        return writer.sheets[safe_sheet]
    ws = workbook.create_sheet(title=safe_sheet)
    writer.sheets[safe_sheet] = ws
    return ws


def _cg_v2_style_header_row(ws, header_row: int, num_cols: int) -> None:
    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF")
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)


def _cg_v2_style_title_row(ws, row_idx: int) -> None:
    cell = ws.cell(row=row_idx, column=1)
    cell.font = Font(bold=True, size=12)
    cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")


def _cg_v2_write_titled_table(
    writer: pd.ExcelWriter,
    sheet_name: str,
    title: str,
    df: pd.DataFrame,
    title_row: int,
    *,
    add_filter: bool = False,
) -> tuple[int, int, int]:
    ws = _cg_v2_get_or_create_sheet(writer, sheet_name)
    ws.cell(row=title_row, column=1, value=title)
    _cg_v2_style_title_row(ws, title_row)
    startrow = title_row
    export_df = df.copy()
    export_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=startrow)
    header_row = title_row + 1
    end_row = header_row + len(export_df)
    _cg_v2_style_header_row(ws, header_row, len(export_df.columns))
    if add_filter and export_df.columns.size > 0 and len(export_df) >= 0:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(export_df.columns))}{max(end_row, header_row)}"
    return title_row, header_row, end_row


def _cg_v2_autosize_worksheet(ws, *, min_width: int = 10, max_width: int = 38) -> None:
    for column_cells in ws.columns:
        col_idx = column_cells[0].column
        lengths: list[int] = []
        for cell in column_cells[:400]:
            if cell.value is not None:
                lengths.append(len(str(cell.value)))
        width = min(max(min_width, (max(lengths) + 2) if lengths else min_width), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_control_gestion_v2_filename(*, semana_inicio: str, vista: str, foco: str | None) -> str:
    semana_token = _safe_file_token(str(semana_inicio or "")[:10], fallback="semana")
    vista_token = _safe_file_token(str(vista or "scope").lower(), fallback="scope")
    foco_token = _safe_file_token(foco, fallback="scope")
    return f"CONTROL_GESTION_V2_{semana_token}_{vista_token}_{foco_token}.xlsx"


def build_control_gestion_v2_workbook(
    *,
    context: dict[str, Any],
    summary: dict[str, Any],
    detail_df: pd.DataFrame,
) -> bytes:
    detail_export_df = _cg_v2_prepare_detail_export_df(detail_df)
    context_df = _cg_v2_context_export_df(context)
    summary_df = _cg_v2_summary_export_df(summary)
    top_clientes_df = _cg_v2_top_incumplimientos_por_cliente(detail_export_df)
    top_pendientes_df = _cg_v2_top_locales_rutas_pendiente(detail_export_df)

    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        resumen_sheet = "Resumen ejecutivo"
        detalle_sheet = "Detalle operativo"

        _, _, resumen_context_end = _cg_v2_write_titled_table(
            writer,
            resumen_sheet,
            "Contexto exportado",
            context_df,
            1,
        )
        _, _, resumen_kpi_end = _cg_v2_write_titled_table(
            writer,
            resumen_sheet,
            "KPIs semanales",
            summary_df,
            resumen_context_end + 2,
        )
        _, _, resumen_top_client_end = _cg_v2_write_titled_table(
            writer,
            resumen_sheet,
            "Top incumplimientos por cliente",
            top_clientes_df,
            resumen_kpi_end + 2,
        )
        _cg_v2_write_titled_table(
            writer,
            resumen_sheet,
            "Top locales/rutas con mayor pendiente",
            top_pendientes_df,
            resumen_top_client_end + 2,
        )

        _, _, detalle_context_end = _cg_v2_write_titled_table(
            writer,
            detalle_sheet,
            "Contexto exportado",
            context_df,
            1,
        )
        _, detalle_header_row, _ = _cg_v2_write_titled_table(
            writer,
            detalle_sheet,
            "Detalle operativo",
            detail_export_df,
            detalle_context_end + 2,
            add_filter=True,
        )

        resumen_ws = writer.sheets[resumen_sheet]
        detalle_ws = writer.sheets[detalle_sheet]
        resumen_ws.freeze_panes = "A2"
        detalle_ws.freeze_panes = f"A{detalle_header_row + 1}"

        for ws in [resumen_ws, detalle_ws]:
            _cg_v2_autosize_worksheet(ws)

        workbook = writer.book
        if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 2:
            del workbook["Sheet"]

    return out.getvalue()


def export_excel_generic(sheet_name: str, df_export: pd.DataFrame) -> bytes:
    out = io.BytesIO()
    safe_sheet = str(sheet_name or "DATA")[:31]
    df_export = _coalesce_duplicate_rr_columns(df_export)
    df_export = _sorted_for_export(df_export)

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name=safe_sheet)
        ws = writer.sheets[safe_sheet]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for col_cells in ws.columns:
            col_letter = col_cells[0].column_letter
            values = []
            for cell in col_cells[:200]:
                if cell.value is not None:
                    values.append(len(str(cell.value)))
            max_len = max(values) if values else 10
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)

    return out.getvalue()


def export_excel_one_sheet(cod_rt: str, df_export: pd.DataFrame) -> bytes:
    return export_excel_generic(str(cod_rt), df_export)


def export_pdf_table(title_lines: list[str], df_export: pd.DataFrame) -> bytes:
    out = io.BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=landscape(A4),
        leftMargin=1.0 * cm,
        rightMargin=1.0 * cm,
        topMargin=1.0 * cm,
        bottomMargin=1.0 * cm,
    )

    styles = getSampleStyleSheet()
    body = ParagraphStyle(
        "BodySmall",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=7,
        leading=8,
    )

    story = []
    for line in title_lines:
        story.append(Paragraph(line, styles["Heading4"]))
    story.append(Spacer(1, 8))

    df = _sorted_for_export(df_export)
    if "Descripción del Producto" not in df.columns:
        df["Descripción del Producto"] = ""
    df["Descripción del Producto"] = df["Descripción del Producto"].astype(str)

    data = [EXPORT_COLS]
    for _, r in df.iterrows():
        row = []
        for c in EXPORT_COLS:
            v = r.get(c, "")
            if c == "Descripción del Producto":
                row.append(Paragraph(str(v), body))
            else:
                row.append(str(v))
        data.append(row)

    col_widths = [
        2.4 * cm,
        2.6 * cm,
        2.6 * cm,
        9.8 * cm,
        1.8 * cm,
        2.0 * cm,
        2.0 * cm,
        3.0 * cm,
        2.0 * cm,
    ]

    table = LongTable(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E6E6E6")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CFCFCF")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F7F7")]),
        ("ALIGN", (4, 1), (4, -1), "RIGHT"),
        ("ALIGN", (5, 1), (7, -1), "CENTER"),
        ("ALIGN", (8, 1), (8, -1), "LEFT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))

    story.append(table)
    doc.build(story)
    return out.getvalue()



def _pdf_column_widths(columns: list[str]) -> tuple[list[float], object]:
    width_map_cm = {
        "Fecha stock": 2.1,
        "FECHA STOCK": 2.1,
        "MARCA": 2.2,
        "Sku": 2.0,
        "SKU": 2.0,
        "COD_RT": 2.0,
        "LOCAL": 4.2,
        "CLIENTE": 3.2,
        "RESP. TIPO": 2.5,
        "RESPONSABLE": 3.2,
        "RUTERO": 3.0,
        "REPONEDOR": 3.4,
        "CLIENTES": 2.2,
        "LOCALES": 2.0,
        "TOTAL SKUS": 2.3,
        "SKUS EN FOCO": 2.5,
        "Stock": 1.6,
        "STOCK": 1.6,
        "VENTA(+7)": 2.0,
        "VENTA 0": 1.9,
        "NEGATIVO": 2.0,
        "RIESGO DE QUIEBRE": 3.0,
        "QUIEBRES OBS.": 2.8,
        "OTROS": 2.4,
        "Descripción del Producto": 8.4,
        "PRODUCTO": 8.4,
        "FOCO PRINCIPAL": 2.8,
        "ACCIÓN SUGERIDA": 4.8,
    }
    widths_cm = [width_map_cm.get(col, 2.5) for col in columns]
    total_cm = sum(widths_cm)
    pagesize = landscape(A3) if len(columns) >= 10 or total_cm > 27.2 else landscape(A4)
    return [w * cm for w in widths_cm], pagesize


def export_pdf_generic(title_lines: list[str], df_export: pd.DataFrame, columns: list[str]) -> bytes:
    out = io.BytesIO()
    col_widths, pagesize = _pdf_column_widths(columns)
    doc = SimpleDocTemplate(
        out,
        pagesize=pagesize,
        leftMargin=0.8 * cm,
        rightMargin=0.8 * cm,
        topMargin=0.8 * cm,
        bottomMargin=0.8 * cm,
    )

    styles = getSampleStyleSheet()
    body = ParagraphStyle(
        "BodySmallWrap",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=6.3,
        leading=7.2,
    )
    header = ParagraphStyle(
        "HeaderSmallWrap",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=7,
        leading=8,
    )

    wrap_cols = {
        "Descripción del Producto",
        "PRODUCTO",
        "ACCIÓN SUGERIDA",
        "OTROS",
        "LOCAL",
        "CLIENTE",
        "RESPONSABLE",
        "RESP. TIPO",
        "RUTERO",
        "REPONEDOR",
        "FOCO PRINCIPAL",
    }
    numeric_cols = {
        "Stock",
        "STOCK",
        "TOTAL SKUS",
        "SKUS EN FOCO",
        "VENTA(+7)",
        "VENTA 0",
        "CLIENTES",
        "LOCALES",
    }
    center_cols = {
        "VENTA 0",
        "NEGATIVO",
        "RIESGO DE QUIEBRE",
        "QUIEBRES OBS.",
        "FOCO PRINCIPAL",
    }

    story = []
    for line in title_lines:
        story.append(Paragraph(line, styles["Heading4"]))
    story.append(Spacer(1, 8))

    df = _sorted_for_export(df_export)
    for c in columns:
        if c not in df.columns:
            df[c] = ""
    df = df[columns].copy()

    data = [[Paragraph(str(col), header) for col in columns]]
    for _, r in df.iterrows():
        row = []
        for c in columns:
            v = r.get(c, "")
            if c in wrap_cols:
                row.append(Paragraph(str(v), body))
            else:
                row.append(str(v))
        data.append(row)

    table = LongTable(data, colWidths=col_widths, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E6E6E6")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CFCFCF")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F7F7")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
    ]

    for idx, col in enumerate(columns):
        if col in numeric_cols:
            style.append(("ALIGN", (idx, 1), (idx, -1), "RIGHT"))
        elif col in center_cols:
            style.append(("ALIGN", (idx, 1), (idx, -1), "CENTER"))
        else:
            style.append(("ALIGN", (idx, 1), (idx, -1), "LEFT"))

    table.setStyle(TableStyle(style))
    story.append(table)
    doc.build(story)
    return out.getvalue()


def export_pdf_focus_table(title_lines: list[str], df_export: pd.DataFrame) -> bytes:
    return export_pdf_generic(title_lines, df_export, FOCUS_EXPORT_COLS)
